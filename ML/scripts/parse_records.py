#!/usr/bin/env python3
"""
Parse Records.xlsx (OneDrive Documents version) into structured CSVs.

Source: /tmp/Records_OneDrive_Docs.xlsx (769K, Jun 2025)
Sheets parsed:
  - Lab (169 tests × 318 dates, 2016-04-28 to 2025-02-03)
  - FoodLog (2021), FoodLog2022, FoodLog2023, FoodLog2024, FoodLog2025
  - Food (old 2016-2018 time-slot format)
  - Poop Log (2022-06-01 to 2025-05-31, 2113 rows)
  - Machine (dialysis machine settings, 56 rows × 165 cols)
  - Home Health Visits
  - Physicians
  - Bills

Output: data/raw/excel/records_*.csv
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime, time as dtime
from pathlib import Path

_ML = Path(__file__).resolve().parent.parent
SRC = str(Path.home() / "Desktop" / "RecordsN.xlsx")
OUT_DIR = str(_ML / "data" / "raw" / "excel")
os.makedirs(OUT_DIR, exist_ok=True)


def parse_lab_sheet(wb_path):
    """Parse the Lab sheet: rows=lab tests, columns=dates. Transpose to long format."""
    print("\n=== Parsing Lab sheet ===")
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb['Lab']
    
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    # Row 0: year headers (sparse), Row 1: dates, Row 2: metric headers
    # Actual lab data starts at row 3
    date_row = all_rows[1]  # row index 1 has dates
    
    # Extract dates from columns (skip first 7 cols which are labels/metadata)
    date_cols = {}
    for ci, val in enumerate(date_row):
        if hasattr(val, 'year'):
            date_cols[ci] = val
    
    print(f"  Found {len(date_cols)} date columns")
    
    # Extract lab test names from column A (rows 3 onward)
    # Row 2 has "Metric", "Reading", "Clinical Pref" headers
    records = []
    for ri in range(3, len(all_rows)):
        row = all_rows[ri]
        test_name = row[0] if row[0] else None
        if not test_name or not isinstance(test_name, str):
            continue
        test_name = test_name.strip()
        if not test_name:
            continue
        
        # Get unit from col B if present
        unit = row[1] if len(row) > 1 and row[1] else None
        # Get clinical pref from col C if present
        clin_pref = row[3] if len(row) > 3 and row[3] else None
        
        # Iterate over date columns
        for ci, dt in date_cols.items():
            if ci < len(row):
                val = row[ci]
                if val is not None and val != '' and val != ' ':
                    # Get lab facility from row index 2 (Lab row)
                    facility = all_rows[2][ci] if ci < len(all_rows[2]) else None
                    records.append({
                        'date': dt.date() if hasattr(dt, 'date') else dt,
                        'test_name': test_name,
                        'value': val,
                        'unit': unit,
                        'clinical_pref': clin_pref,
                        'facility': facility,
                    })
    
    df = pd.DataFrame(records)
    
    # Try to convert values to numeric where possible
    df['value_numeric'] = pd.to_numeric(df['value'], errors='coerce')
    df['value_text'] = df['value'].where(df['value_numeric'].isna()).astype(str).replace('nan', '')
    
    # Sort by date and test
    df = df.sort_values(['date', 'test_name']).reset_index(drop=True)
    
    out = os.path.join(OUT_DIR, "records_lab.csv")
    df.to_csv(out, index=False)
    print(f"  Saved {len(df)} lab results to {out}")
    print(f"  Date range: {df['date'].min()} to {df['date'].max()}")
    print(f"  Unique tests: {df['test_name'].nunique()}")
    print(f"  Unique dates: {df['date'].nunique()}")
    return df


def parse_food_old(wb_path):
    """Parse the old Food sheet (2016-2018, time-slot layout)."""
    print("\n=== Parsing Food sheet (old, 2016-2018) ===")
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb['Food']
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    # Row 0: time slot headers
    time_slots = all_rows[0][1:]  # skip col A (date)
    
    records = []
    for ri in range(1, len(all_rows)):
        row = all_rows[ri]
        dt = row[0]
        if not hasattr(dt, 'year'):
            continue
        for ci, slot in enumerate(time_slots, start=1):
            if ci < len(row) and row[ci] and str(row[ci]).strip():
                records.append({
                    'date': dt.date(),
                    'time_slot': slot,
                    'description': str(row[ci]).strip(),
                    'source': 'Food_old',
                })
    
    df = pd.DataFrame(records)
    print(f"  {len(df)} food entries from old format")
    return df


def parse_foodlog(wb_path, sheet_name, year_hint=None):
    """Parse a FoodLog sheet (2021-2025 format). Each sheet has slightly different column layout."""
    print(f"\n=== Parsing {sheet_name} ===")
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    if not all_rows:
        return pd.DataFrame()
    
    # Detect header row — look for 'Date' or 'Description' in first few rows
    header_idx = 0
    for i, row in enumerate(all_rows[:3]):
        vals = [str(v).strip().lower() if v else '' for v in row]
        if 'date' in vals or 'description' in vals:
            header_idx = i
            break
    
    headers_raw = all_rows[header_idx]
    
    # Map column indices to standard names (first match wins to avoid duplicates)
    col_map = {}
    for ci, h in enumerate(headers_raw):
        if h is None:
            continue
        h_lower = str(h).strip().lower()
        key = None
        if h_lower == 'date':
            key = 'date'
        elif 'start' in h_lower and 'time' in h_lower:
            key = 'start_time'
        elif 'end' in h_lower and 'time' in h_lower:
            key = 'end_time'
        elif h_lower == 'description':
            key = 'description'
        elif h_lower == 'quantity':
            key = 'quantity'
        elif 'pre' in h_lower and 'weight' in h_lower:
            key = 'pre_weight'
        elif h_lower == 'pre' and 'pre_weight' not in col_map:
            key = 'pre_weight'
        elif 'post' in h_lower and 'weight' in h_lower:
            key = 'post_weight'
        elif h_lower == 'post' and 'post_weight' not in col_map:
            key = 'post_weight'
        elif 'total' in h_lower and 'weight' in h_lower:
            key = 'total_weight'
        elif 'state' in h_lower:
            key = 'state_post'
        elif 'medication' in h_lower or 'med' in h_lower.split():
            key = 'medication'
        elif h_lower == 'gain':
            key = 'gain'
        elif 'image' in h_lower:
            key = 'image_url'
        elif 'new recipe' in h_lower or h_lower == 'new recipe':
            key = 'new_recipe'
        elif 'recipe code' in h_lower or 'redipe code' in h_lower:
            key = 'recipe_code'
        elif 'recipe item' in h_lower:
            key = 'recipe_ration'
        elif 'recipe prep' in h_lower:
            key = 'recipe_prep'
        elif 'recipe nutrient' in h_lower:
            key = 'recipe_nutrient_code'
        elif 'recipe metabolic' in h_lower:
            key = 'recipe_metabolic'
        
        # First match wins — skip if already mapped
        if key and key not in col_map:
            col_map[key] = ci
    
    # Special case: FoodLog (2021) has date in col H (index 7)
    if sheet_name == 'FoodLog' and 'date' not in col_map:
        col_map['date'] = 7
        col_map['start_time'] = 8
        col_map['end_time'] = 9
        col_map['image_url'] = 10
        col_map['description'] = 11
        col_map['quantity'] = 12
        col_map['pre_weight'] = 13
        col_map['post_weight'] = 14
        col_map['state_post'] = 15
    
    print(f"  Column mapping: {col_map}")
    
    date_ci = col_map.get('date', 0)
    
    records = []
    for ri in range(header_idx + 1, len(all_rows)):
        row = all_rows[ri]
        if len(row) <= date_ci:
            continue
        dt = row[date_ci]
        if not hasattr(dt, 'year'):
            continue
        
        def get_val(key):
            ci = col_map.get(key)
            if ci is not None and ci < len(row):
                v = row[ci]
                # Skip formula strings
                if isinstance(v, str) and v.startswith('='):
                    return None
                return v
            return None
        
        desc = get_val('description')
        if desc and isinstance(desc, str) and desc.startswith('='):
            desc = None
        
        rec = {
            'date': dt.date() if hasattr(dt, 'date') else dt,
            'start_time': get_val('start_time'),
            'end_time': get_val('end_time'),
            'description': desc,
            'quantity': get_val('quantity'),
            'pre_weight_kg': get_val('pre_weight'),
            'post_weight_kg': get_val('post_weight'),
            'state_post': get_val('state_post'),
            'medication': get_val('medication'),
            'new_recipe': get_val('new_recipe'),
            'recipe_code': get_val('recipe_code'),
            'source': sheet_name,
        }
        
        # Convert time objects
        for tk in ['start_time', 'end_time']:
            v = rec[tk]
            if isinstance(v, dtime):
                rec[tk] = v.strftime('%H:%M')
            elif hasattr(v, 'strftime'):
                rec[tk] = v.strftime('%H:%M')
            elif v is not None:
                rec[tk] = str(v)
        
        records.append(rec)
    
    df = pd.DataFrame(records)
    
    # Clean weights
    for wc in ['pre_weight_kg', 'post_weight_kg']:
        if wc in df.columns:
            df[wc] = pd.to_numeric(df[wc], errors='coerce')
            # Filter impossible weights
            df.loc[df[wc] < 20, wc] = np.nan
            df.loc[df[wc] > 150, wc] = np.nan
    
    print(f"  {len(df)} food entries")
    if not df.empty:
        print(f"  Date range: {df['date'].min()} to {df['date'].max()}")
    return df


def parse_poop_log(wb_path):
    """Parse the Poop Log sheet."""
    print("\n=== Parsing Poop Log ===")
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb['Poop Log']
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    # Header: Date, Time, Last Meal, Blood?, Notes
    records = []
    for ri in range(1, len(all_rows)):
        row = all_rows[ri]
        dt = row[0]
        if not hasattr(dt, 'year'):
            continue
        
        tm = row[1] if len(row) > 1 else None
        last_meal = row[2] if len(row) > 2 else None
        blood = row[3] if len(row) > 3 else None
        notes = row[4] if len(row) > 4 else None
        
        # Convert time
        if isinstance(tm, dtime):
            tm = tm.strftime('%H:%M')
        elif hasattr(tm, 'strftime'):
            tm = tm.strftime('%H:%M')
        
        # Convert last_meal time
        if isinstance(last_meal, dtime):
            last_meal = last_meal.strftime('%H:%M')
        elif hasattr(last_meal, 'strftime'):
            last_meal = last_meal.strftime('%H:%M')
        elif isinstance(last_meal, str) and last_meal.startswith('='):
            last_meal = None
        
        # Normalize blood field
        if blood and isinstance(blood, str):
            blood = blood.strip().upper()
            if blood in ('YES', 'Y'):
                blood = True
            elif blood in ('NO', 'N', 'NONE'):
                blood = False
            else:
                blood = blood  # keep as-is for unusual values
        
        records.append({
            'date': dt.date(),
            'time': tm,
            'last_meal_time': last_meal,
            'blood': blood,
            'notes': notes,
        })
    
    df = pd.DataFrame(records)
    print(f"  {len(df)} poop log entries")
    print(f"  Date range: {df['date'].min()} to {df['date'].max()}")
    blood_count = df['blood'].sum() if df['blood'].dtype == bool else (df['blood'] == True).sum()
    print(f"  Blood present: {blood_count} entries")
    return df


def parse_machine(wb_path):
    """Parse the Machine sheet (dialysis machine settings over time)."""
    print("\n=== Parsing Machine sheet ===")
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb['Machine']
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    # Row 1 has dates starting from col 7
    date_row = all_rows[1]
    date_cols = {}
    for ci, val in enumerate(date_row):
        if hasattr(val, 'year'):
            date_cols[ci] = val
    
    print(f"  Found {len(date_cols)} date columns")
    
    # Rows 2+ have metric name in col 0, unit in col 1
    records = []
    for ri in range(2, len(all_rows)):
        row = all_rows[ri]
        metric = row[0] if row[0] else None
        if not metric or not isinstance(metric, str):
            continue
        metric = metric.strip()
        if not metric:
            continue
        
        unit = row[1] if len(row) > 1 and row[1] else None
        
        for ci, dt in date_cols.items():
            if ci < len(row):
                val = row[ci]
                if val is not None and val != '' and val != ' ':
                    if isinstance(val, str) and val.startswith('='):
                        continue
                    records.append({
                        'date': dt.date(),
                        'metric': metric,
                        'value': val,
                        'unit': unit,
                    })
    
    df = pd.DataFrame(records)
    df['value_numeric'] = pd.to_numeric(df['value'], errors='coerce')
    df = df.sort_values(['date', 'metric']).reset_index(drop=True)
    
    print(f"  {len(df)} machine records")
    if not df.empty:
        print(f"  Date range: {df['date'].min()} to {df['date'].max()}")
        print(f"  Unique metrics: {df['metric'].nunique()}")
    return df


def parse_physicians(wb_path):
    """Parse the Physicians sheet."""
    print("\n=== Parsing Physicians ===")
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb['Physicians']
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    # Row 1: headers (FullName, Facility, Specialty, DFC, DFS)
    records = []
    for ri in range(2, len(all_rows)):
        row = all_rows[ri]
        if len(row) < 5 or not row[1]:
            continue
        records.append({
            'full_name': row[1],
            'facility': row[2],
            'specialty': row[3],
            'date_first_contact': row[4].date() if hasattr(row[4], 'date') else row[4],
            'date_first_seen': row[5].date() if len(row) > 5 and hasattr(row[5], 'date') else (row[5] if len(row) > 5 else None),
        })
    
    df = pd.DataFrame(records)
    print(f"  {len(df)} physicians")
    return df


def parse_home_health(wb_path):
    """Parse Home Health Visits sheet."""
    print("\n=== Parsing Home Health Visits ===")
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb['Home Health Visits']
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    # Row 1 has dates starting at col 5
    date_row = all_rows[1]
    date_cols = {}
    for ci, val in enumerate(date_row):
        if hasattr(val, 'year'):
            date_cols[ci] = val
    
    records = []
    for ri in range(2, len(all_rows)):
        row = all_rows[ri]
        if not row[1]:
            continue
        referer = row[1]
        visit_id = row[2]
        visit_type = row[3]
        quantity = row[4]
        
        # Each date column has the visit time/notes
        for ci, dt in date_cols.items():
            if ci < len(row) and row[ci]:
                records.append({
                    'date': dt.date(),
                    'referer': referer,
                    'visit_id': visit_id,
                    'visit_type': visit_type,
                    'quantity_prescribed': quantity,
                    'notes': row[ci],
                })
    
    df = pd.DataFrame(records)
    print(f"  {len(df)} home health visit records")
    return df


def main():
    print(f"Source: {SRC}")
    print(f"Output: {OUT_DIR}")
    
    # 1. Lab results
    lab_df = parse_lab_sheet(SRC)
    
    # 2. Food logs — combine all sheets
    food_dfs = []
    
    # Old format (2016-2018)
    food_old = parse_food_old(SRC)
    if not food_old.empty:
        # Old format already has 'description' from parsing; just append
        food_dfs.append(food_old)
    
    # FoodLog sheets
    for sheet in ['FoodLog', 'FoodLog2022', 'FoodLog2023', 'FoodLog2024', 'FoodLog2025']:
        df = parse_foodlog(SRC, sheet)
        if not df.empty:
            food_dfs.append(df)
    
    if food_dfs:
        food_all = pd.concat(food_dfs, ignore_index=True)
        food_all = food_all.sort_values(['date', 'start_time']).reset_index(drop=True)
        out = os.path.join(OUT_DIR, "records_food.csv")
        food_all.to_csv(out, index=False)
        print(f"\n=== FOOD TOTAL: {len(food_all)} entries, {food_all['date'].nunique()} unique dates ===")
        print(f"  Date range: {food_all['date'].min()} to {food_all['date'].max()}")
        print(f"  Saved to {out}")
    
    # 3. Poop log
    poop_df = parse_poop_log(SRC)
    out = os.path.join(OUT_DIR, "records_poop.csv")
    poop_df.to_csv(out, index=False)
    
    # 4. Machine settings
    machine_df = parse_machine(SRC)
    out = os.path.join(OUT_DIR, "records_machine.csv")
    machine_df.to_csv(out, index=False)
    
    # 5. Physicians
    phys_df = parse_physicians(SRC)
    out = os.path.join(OUT_DIR, "records_physicians.csv")
    phys_df.to_csv(out, index=False)
    
    # 6. Home Health Visits
    hh_df = parse_home_health(SRC)
    out = os.path.join(OUT_DIR, "records_home_health.csv")
    hh_df.to_csv(out, index=False)
    
    # Summary
    print("\n" + "=" * 60)
    print("  RECORDS PARSING COMPLETE")
    print("=" * 60)
    print(f"  Lab results:        {len(lab_df):>6} rows ({lab_df['test_name'].nunique()} tests, {lab_df['date'].nunique()} dates)")
    total_food = sum(len(d) for d in food_dfs) if food_dfs else 0
    print(f"  Food log:           {total_food:>6} entries")
    print(f"  Poop log:           {len(poop_df):>6} entries")
    print(f"  Machine settings:   {len(machine_df):>6} records")
    print(f"  Physicians:         {len(phys_df):>6} records")
    print(f"  Home health visits: {len(hh_df):>6} records")
    print(f"\n  Output directory: {OUT_DIR}")


if __name__ == "__main__":
    main()
