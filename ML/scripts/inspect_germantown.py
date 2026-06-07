#!/usr/bin/env python3
"""Quick inspect of FlowsheetGermantown.xlsx using openpyxl read_only mode."""
import os
import re
from pathlib import Path
from openpyxl import load_workbook

FPATH = os.environ.get('GERMANTOWN_XLSX', str(Path(__file__).resolve().parent.parent / 'data' / 'raw' / 'FlowsheetGermantown.xlsx'))

print("Loading workbook (read_only)...")
wb = load_workbook(FPATH, read_only=True, data_only=True)
sheets = wb.sheetnames
print(f"Total sheets: {len(sheets)}")

SKIP = {'VomitLog2021','VomitLog2022','Summary','Maintenance Log',
        'Cleaning Log','Supplies','Dummy','Master'}
session_sheets = [s for s in sheets if s not in SKIP]
print(f"Session sheets: {len(session_sheets)}")

# Year distribution
year_counts = {}
for s in session_sheets:
    m = re.search(r'20\d{2}', s)
    if m:
        yr = int(m.group())
        year_counts[yr] = year_counts.get(yr, 0) + 1
    else:
        year_counts['unknown'] = year_counts.get('unknown', 0) + 1

print("\nYear breakdown:")
for yr in sorted(k for k in year_counts if isinstance(k, int)):
    print(f"  {yr}: {year_counts[yr]} sheets")
if 'unknown' in year_counts:
    print(f"  unknown: {year_counts['unknown']} sheets")
    unknown = [s for s in session_sheets if not re.search(r'20\d{2}', s)]
    print(f"    examples: {unknown[:10]}")

print(f"\nFirst 5: {session_sheets[:5]}")
print(f"Last 5:  {session_sheets[-5:]}")

# Peek at a 2022 sheet
for s in session_sheets:
    if '2022' in s and 'Vomit' not in s:
        ws = wb[s]
        print(f"\nSample sheet: '{s}'")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, max_col=12, values_only=True)):
            vals = [str(v)[:25] if v is not None else '.' for v in row]
            print(f"  R{i:02d}: {'|'.join(vals)}")
        break

# Peek at a 2019 sheet
for s in session_sheets:
    if '2019' in s:
        ws = wb[s]
        print(f"\nSample sheet: '{s}'")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, max_col=12, values_only=True)):
            vals = [str(v)[:25] if v is not None else '.' for v in row]
            print(f"  R{i:02d}: {'|'.join(vals)}")
        break

wb.close()
