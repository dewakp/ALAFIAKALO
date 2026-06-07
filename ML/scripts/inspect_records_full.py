#!/usr/bin/env python3
"""Inspect OneDrive Documents/Records.xlsx - the most complete version."""
import openpyxl

wb = openpyxl.load_workbook("/tmp/Records_OneDrive_Docs.xlsx", read_only=True)
print(f"Sheets: {wb.sheetnames}\n")

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"=== [{sname}] max_row={ws.max_row}, max_col={ws.max_column} ===")
    
    # Print first 3 rows (headers)
    rows = list(ws.iter_rows(max_row=3, values_only=True))
    for i, r in enumerate(rows):
        # Truncate long rows
        display = str(r)
        if len(display) > 300:
            display = display[:300] + "..."
        print(f"  row{i}: {display}")
    
    # Find date range in col A (or row 0 for Lab)
    if sname == 'Lab':
        row0 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        dates = [c for c in row0 if hasattr(c, 'year')]
        if dates:
            print(f"  >> {len(dates)} date columns: {min(dates).date()} to {max(dates).date()}")
    else:
        first_date = None
        last_date = None
        count = 0
        for r in ws.iter_rows(values_only=True):
            if r[0] and hasattr(r[0], 'year'):
                if first_date is None:
                    first_date = r[0]
                last_date = r[0]
                count += 1
        if count > 0:
            print(f"  >> {count} dated rows: {first_date.date()} to {last_date.date()}")
    
    # Last 3 non-empty rows
    all_rows = list(ws.iter_rows(values_only=True))
    non_empty = [(i, r) for i, r in enumerate(all_rows) if any(c is not None for c in r)]
    if len(non_empty) > 3:
        print(f"  Last 3 non-empty rows:")
        for idx, r in non_empty[-3:]:
            display = str(r)
            if len(display) > 300:
                display = display[:300] + "..."
            print(f"    row{idx}: {display}")
    print()

wb.close()
