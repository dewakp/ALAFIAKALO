#!/usr/bin/env python3
"""Inspect Records Excel files to understand structure and date ranges."""

import openpyxl
import sys

from pathlib import Path
import os

_ML = Path(__file__).resolve().parent.parent

files = [
    ("OneDrive 6IGMA_H (696K, Nov 2024)", "/tmp/Records_OneDrive.xlsx"),
    ("local data/Records.xlsx (696K, Nov 2024)", os.environ.get('LOCAL_RECORDS', str(_ML / 'data' / 'raw' / 'excel' / 'Records.xlsx'))),
    ("local data/RecordsData.xlsx (547K, Aug 2024)", os.environ.get('LOCAL_RECORDS_DATA', str(_ML / 'data' / 'raw' / 'excel' / 'RecordsData.xlsx'))),
]

for label, path in files:
    print(f"\n{'='*60}")
    print(f"  {label}")
    print(f"{'='*60}")
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception as e:
        print(f"  ERROR: {e}")
        continue
    
    print(f"  Sheets: {wb.sheetnames}")
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n  --- [{sname}] max_row={ws.max_row}, max_col={ws.max_column} ---")
        # Print first 5 rows
        for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
            print(f"    row{i}: {row}")
        # Print last 5 rows to see date range
        if ws.max_row and ws.max_row > 10:
            print(f"    ... (skipping to last 5 rows)")
            all_rows = list(ws.iter_rows(min_row=max(1, ws.max_row - 4), values_only=True))
            for i, row in enumerate(all_rows):
                rnum = ws.max_row - len(all_rows) + i + 1
                print(f"    row{rnum}: {row}")
    wb.close()
