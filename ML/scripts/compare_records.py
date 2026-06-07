#!/usr/bin/env python3
"""Compare OneDrive vs local Records.xlsx and check date ranges."""
import os
import openpyxl
import hashlib
from pathlib import Path

_ML = Path(__file__).resolve().parent.parent

# Compare checksums
print("=== CHECKSUM COMPARISON ===")
for label, path in [
    ("OneDrive 6IGMA_H", "/tmp/Records_OneDrive.xlsx"),
    ("local data/", os.environ.get('LOCAL_RECORDS', str(_ML / 'data' / 'raw' / 'excel' / 'Records.xlsx'))),
]:
    h = hashlib.md5(open(path, 'rb').read()).hexdigest()
    print(f"  {label}: {h}")

# Inspect OneDrive copy
wb = openpyxl.load_workbook("/tmp/Records_OneDrive.xlsx", read_only=True)
print(f"\n=== SHEETS: {wb.sheetnames} ===")

# Check Lab sheet - dates are in row 0 (columns)
ws = wb['Lab']
row0 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
dates = [c for c in row0 if hasattr(c, 'year')]
if dates:
    print(f"\nLab: {len(dates)} date columns, range {min(dates).date()} to {max(dates).date()}")

# Check each FoodLog sheet
for sname in wb.sheetnames:
    if 'Food' in sname or 'Poop' in sname or 'Dialysis' in sname:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        first_date = None
        last_date = None
        count = 0
        for r in rows:
            if r[0] and hasattr(r[0], 'year'):
                if first_date is None:
                    first_date = r[0]
                last_date = r[0]
                count += 1
        print(f"{sname}: {count} dated rows, {first_date} to {last_date}")

# Check for any 2025 content
for sname in wb.sheetnames:
    if '2025' in sname.lower():
        print(f"\nFound 2025 sheet: {sname}")

wb.close()
