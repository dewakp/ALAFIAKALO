#!/usr/bin/env python3
"""Examine FlowsheetGermantown.xlsx structure."""
import openpyxl

path = "/Users/woleakpose/Desktop/FlowsheetGermantown.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
sheets = wb.sheetnames
print(f"SHEETS ({len(sheets)}):")
for i, s in enumerate(sheets):
    print(f"  {i+1}. {s}")
print()

# Identify session sheets vs utility sheets
skip = {'Sheet1','Sheet2','Sheet3','Summary','VomitLog','AccessMaintenance',
        'Supplies Request','Vomit Log','VomitLog2023','Access Maintenance',
        'Template','template'}
session_sheets = [s for s in sheets if s not in skip]
print(f"SESSION SHEETS: {len(session_sheets)}")
if session_sheets:
    print(f"  First: {session_sheets[0]}")
    print(f"  Last:  {session_sheets[-1]}")
print()

# Examine the first session sheet in detail
if session_sheets:
    ws = wb[session_sheets[0]]
    print(f"=== EXAMINING: '{session_sheets[0]}' ===")
    print(f"  Max rows: {ws.max_row}, Max cols: {ws.max_column}")
    print()
    print("ALL CELLS WITH DATA (rows 1-60):")
    for row in ws.iter_rows(max_row=60, values_only=False):
        vals = [(f"{c.column_letter}{c.row}", c.value) for c in row if c.value is not None]
        if vals:
            print(f"  {vals}")

    print()
    # Check a second session sheet to compare
    if len(session_sheets) > 5:
        ws2 = wb[session_sheets[5]]
        print(f"=== EXAMINING: '{session_sheets[5]}' ===")
        for row in ws2.iter_rows(max_row=55, values_only=False):
            vals = [(f"{c.column_letter}{c.row}", c.value) for c in row if c.value is not None]
            if vals:
                print(f"  {vals}")

# Check for non-session sheets
non_session = [s for s in sheets if s in skip]
print(f"\nNON-SESSION SHEETS: {non_session}")
for ns in non_session:
    ws = wb[ns]
    print(f"\n  '{ns}': max_row={ws.max_row}, max_col={ws.max_column}")
    # Print first 5 rows
    for i, row in enumerate(ws.iter_rows(max_row=5, values_only=False)):
        vals = [(f"{c.column_letter}{c.row}", c.value) for c in row if c.value is not None]
        if vals:
            print(f"    Row {i+1}: {vals[:10]}")

wb.close()
