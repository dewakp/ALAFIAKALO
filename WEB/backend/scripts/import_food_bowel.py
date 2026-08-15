#!/usr/bin/env python3
"""
Import food logs and bowel movement (poop) logs from Records.xlsx into the database.

Food Log Sheets (4 sheets, different column layouts):
  - FoodLog (2021):     H=Date, I=Start, J=End, L=Description, M=Quantity, N=Pre, O=Post, P=State, R=GAIN
  - FoodLog2022:        A=Date, B=Start, C=End, E=Description, F=Quantity, G=Pre, H=Post, I=State, K=GAIN
  - FoodLog2023:        A=Date, B=Start, C=End, E=Description, F=Quantity, G=Pre, H=Post, I=State, K=GAIN
  - FoodLog2024:        A=Date, B=Start, C=End, E=Description, F=Quantity, G=Pre, H=Post, I=GAIN, N=State, O=Medication

Poop Log Sheet (1 sheet, ~1188 rows):
  - A=Date, B=Time, C=Last Meal Time, D=Blood?, E=Notes

Maps to:
  - nutrition_logs table (log_date, meal_type, food_name, serving_size, notes)
  - bowel_movements table (log_date, log_time, blood_present, notes)
"""

import psycopg2

import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _target import resolve_user, resolve_condition  # noqa: E402
import openpyxl
from datetime import datetime, date, time
import re
import traceback

DB_PARAMS = {
    'host': 'localhost',
    'port': 5432,
    'dbname': 'alafia',
    'user': 'alafia',
    'password': 'alafia'
}

# ── Target ────────────────────────────────────────────────────────────────
# USER_ID was hardcoded to a literal row id that exists in no database here.
# Resolve the patient by EMAIL instead — see scripts/_target.py. Test data
# belongs to developer@hntsolutions.com by convention.
#
#   USER_ID = resolve_user(conn, "developer@hntsolutions.com")
#
# Left as None so an accidental run fails loudly instead of writing clinical
# rows against the wrong patient.
USER_ID = None
EXCEL_PATH = '/Users/woleakpose/Library/Mobile Documents/com~apple~CloudDocs/Records.xlsx'


def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s != '#N/A' else None


def safe_float(val):
    if val is None:
        return None
    try:
        v = float(val)
        return v
    except (ValueError, TypeError):
        return None


def parse_time(val):
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        m = re.match(r'(\d{1,2}):(\d{2})', val)
        if m:
            return time(int(m.group(1)), int(m.group(2)))
    return None


def classify_meal(t):
    """Classify meal type based on time of day."""
    if t is None:
        return 'meal'
    if isinstance(t, datetime):
        t = t.time()
    if not isinstance(t, time):
        return 'meal'
    hour = t.hour
    if 5 <= hour < 11:
        return 'breakfast'
    elif 11 <= hour < 15:
        return 'lunch'
    elif 15 <= hour < 17:
        return 'snack'
    elif 17 <= hour < 22:
        return 'dinner'
    else:
        return 'snack'  # late night or early morning


def import_foodlog_2021(wb, conn, existing_keys):
    """Import FoodLog sheet (2021 data, header row 2, data starts row 3)."""
    ws = wb['FoodLog']
    imported = 0
    cur = conn.cursor()

    for row in ws.iter_rows(min_row=3, values_only=False):
        vals = {}
        for c in row:
            if hasattr(c, 'column_letter') and c.value is not None:
                vals[c.column_letter] = c.value

        meal_date = vals.get('H')
        if not isinstance(meal_date, (datetime, date)):
            continue
        if isinstance(meal_date, datetime):
            meal_date = meal_date.date()

        start_time = parse_time(vals.get('I'))
        description = safe_str(vals.get('L'))
        if not description:
            continue

        quantity = safe_str(vals.get('M'))
        pre_weight = safe_float(vals.get('N'))
        post_weight = safe_float(vals.get('O'))
        state_post = safe_str(vals.get('P'))
        gain = safe_float(vals.get('R'))

        meal_type = classify_meal(start_time)
        dedup_key = (meal_date, description[:100] if description else '')
        if dedup_key in existing_keys:
            continue

        notes_parts = []
        if start_time:
            end_time = parse_time(vals.get('J'))
            notes_parts.append(f"Eaten: {start_time.strftime('%H:%M')}" +
                             (f"-{end_time.strftime('%H:%M')}" if end_time else ""))
        if pre_weight:
            notes_parts.append(f"Pre-weight: {pre_weight}kg")
        if post_weight:
            notes_parts.append(f"Post-weight: {post_weight}kg")
        if gain:
            notes_parts.append(f"Weight gain: {gain}kg")
        if state_post:
            notes_parts.append(f"State: {state_post}")

        notes = '; '.join(notes_parts) if notes_parts else None

        cur.execute("""
            INSERT INTO nutrition_logs (user_id, log_date, meal_type, food_name, serving_size, notes, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (USER_ID, meal_date, meal_type, description, quantity, notes, datetime.utcnow()))
        existing_keys.add(dedup_key)
        imported += 1

    conn.commit()
    return imported


def import_foodlog_standard(wb, conn, sheet_name, existing_keys):
    """Import FoodLog2022/2023/2024 sheets (standard A-based layout)."""
    ws = wb[sheet_name]
    imported = 0
    cur = conn.cursor()

    # Detect header row
    header_row = 1
    first_cell = ws['A1'].value
    if first_cell and isinstance(first_cell, str) and first_cell.strip() != 'Date':
        header_row = 2  # FoodLog2022 has a sub-header in row 1

    # Determine column mapping per sheet
    # FoodLog2022: A=Date, B=Start, C=End, E=Desc, F=Qty, G=Pre, H=Post, I=State, K=GAIN
    # FoodLog2023: A=Date, B=Start, C=End, E=Desc, F=Qty, G=Pre, H=Post, I=State, K=GAIN
    # FoodLog2024: A=Date, B=Start, C=End, E=Desc, F=Qty, G=Pre, H=Post, I=GAIN, N=State, O=Medication
    is_2024 = '2024' in sheet_name

    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        vals = {c.column_letter: c.value for c in row}

        meal_date = vals.get('A')
        if not isinstance(meal_date, (datetime, date)):
            continue
        if isinstance(meal_date, datetime):
            meal_date = meal_date.date()

        start_time = parse_time(vals.get('B'))
        end_time = parse_time(vals.get('C'))
        description = safe_str(vals.get('E'))
        if not description:
            continue

        quantity = safe_str(vals.get('F'))
        pre_weight = safe_float(vals.get('G'))
        post_weight = safe_float(vals.get('H'))

        if is_2024:
            gain = safe_float(vals.get('I'))
            state_post = safe_str(vals.get('N'))
            medication = safe_str(vals.get('O'))
        else:
            state_post = safe_str(vals.get('I'))
            gain = safe_float(vals.get('K'))
            medication = None

        meal_type = classify_meal(start_time)
        dedup_key = (meal_date, description[:100] if description else '')
        if dedup_key in existing_keys:
            continue

        notes_parts = []
        if start_time:
            notes_parts.append(f"Eaten: {start_time.strftime('%H:%M')}" +
                             (f"-{end_time.strftime('%H:%M')}" if end_time else ""))
        if pre_weight:
            notes_parts.append(f"Pre-weight: {pre_weight}kg")
        if post_weight:
            notes_parts.append(f"Post-weight: {post_weight}kg")
        if gain:
            notes_parts.append(f"Weight gain: {gain}kg")
        if state_post and state_post not in ('#N/A',):
            notes_parts.append(f"State: {state_post}")
        if medication:
            notes_parts.append(f"Medication: {medication}")

        notes = '; '.join(notes_parts) if notes_parts else None

        cur.execute("""
            INSERT INTO nutrition_logs (user_id, log_date, meal_type, food_name, serving_size, notes, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (USER_ID, meal_date, meal_type, description, quantity, notes, datetime.utcnow()))
        existing_keys.add(dedup_key)
        imported += 1

    conn.commit()
    return imported


def import_poop_log(wb, conn):
    """Import Poop Log sheet into bowel_movements table."""
    ws = wb['Poop Log']
    imported = 0
    cur = conn.cursor()

    # Get existing to dedup
    cur.execute("SELECT log_date, log_time FROM bowel_movements WHERE user_id = %s", (USER_ID,))
    existing = {(row[0], row[1]) for row in cur.fetchall()}

    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = {c.column_letter: c.value for c in row}

        poop_date = vals.get('A')
        if not isinstance(poop_date, (datetime, date)):
            continue
        if isinstance(poop_date, datetime):
            poop_date = poop_date.date()

        poop_time = parse_time(vals.get('B'))
        if poop_time is None:
            poop_time = time(0, 0)

        dedup_key = (poop_date, poop_time)
        if dedup_key in existing:
            continue

        last_meal = vals.get('C')
        blood_str = safe_str(vals.get('D'))
        blood_present = None
        if blood_str:
            blood_present = blood_str.upper() in ('YES', 'Y', 'TRUE', '1')

        notes_raw = safe_str(vals.get('E'))

        # Additional columns F-I may exist in some rows
        extra_notes = []
        for col in ['F', 'G', 'H', 'I']:
            v = safe_str(vals.get(col))
            if v:
                extra_notes.append(v)

        notes_parts = []
        if last_meal:
            t = parse_time(last_meal)
            if t:
                notes_parts.append(f"Last meal: {t.strftime('%H:%M')}")
        if notes_raw:
            notes_parts.append(notes_raw)
        if extra_notes:
            notes_parts.extend(extra_notes)
        notes = '; '.join(notes_parts) if notes_parts else None

        cur.execute("""
            INSERT INTO bowel_movements (user_id, log_date, log_time, blood_present, notes, created_at)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (USER_ID, poop_date, poop_time, blood_present, notes, datetime.utcnow()))
        existing.add(dedup_key)
        imported += 1

    conn.commit()
    return imported


def main():
    print("=" * 60)
    print("ALAFIA Food & Bowel Movement Import")
    print("=" * 60)

    conn = psycopg2.connect(**DB_PARAMS)
    cur = conn.cursor()

    # Current state
    cur.execute("SELECT COUNT(*) FROM nutrition_logs WHERE user_id = %s", (USER_ID,))
    print(f"  Current nutrition_logs: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM bowel_movements WHERE user_id = %s", (USER_ID,))
    print(f"  Current bowel_movements: {cur.fetchone()[0]}")

    print(f"\nLoading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # Build dedup keys for existing nutrition logs
    cur.execute("SELECT log_date, LEFT(food_name, 100) FROM nutrition_logs WHERE user_id = %s", (USER_ID,))
    existing_food_keys = {(row[0], row[1]) for row in cur.fetchall()}

    # Import Food Logs
    total_food = 0

    print("\n--- FoodLog (2021) ---")
    try:
        n = import_foodlog_2021(wb, conn, existing_food_keys)
        print(f"  Imported: {n} entries")
        total_food += n
    except Exception as e:
        print(f"  ERROR: {e}")
        traceback.print_exc()
        conn.rollback()

    for sheet in ['FoodLog2022', 'FoodLog2023', 'FoodLog2024']:
        print(f"\n--- {sheet} ---")
        try:
            n = import_foodlog_standard(wb, conn, sheet, existing_food_keys)
            print(f"  Imported: {n} entries")
            total_food += n
        except Exception as e:
            print(f"  ERROR: {e}")
            traceback.print_exc()
            conn.rollback()

    # Import Poop Log
    print("\n--- Poop Log ---")
    try:
        poop_count = import_poop_log(wb, conn)
        print(f"  Imported: {poop_count} entries")
    except Exception as e:
        print(f"  ERROR: {e}")
        traceback.print_exc()
        conn.rollback()
        poop_count = 0

    wb.close()

    # Final state
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*), MIN(log_date), MAX(log_date) FROM nutrition_logs WHERE user_id = %s", (USER_ID,))
    food_row = cur.fetchone()
    cur.execute("SELECT COUNT(*), MIN(log_date), MAX(log_date) FROM bowel_movements WHERE user_id = %s", (USER_ID,))
    poop_row = cur.fetchone()

    conn.close()

    print(f"\n{'='*60}")
    print(f"IMPORT COMPLETE")
    print(f"{'='*60}")
    print(f"  Food logs imported: {total_food}")
    print(f"  Bowel movements imported: {poop_count}")
    print(f"\n  Total nutrition_logs: {food_row[0]}  ({food_row[1]} to {food_row[2]})")
    print(f"  Total bowel_movements: {poop_row[0]}  ({poop_row[1]} to {poop_row[2]})")


if __name__ == '__main__':
    main()
