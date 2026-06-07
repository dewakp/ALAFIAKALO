#!/usr/bin/env python3
"""Examine hemodialysis FlowSheet workbooks to understand their structure."""

import openpyxl

files = [
    '/Users/woleakpose/Documents/2023FlowSheets-ex.xlsx',
    '/Users/woleakpose/Documents/Developer/data/2024FlowSheets.xlsx',
    '/Users/woleakpose/Documents/Developer/WellnessScore/Flowsheets.xlsx',
]

for f in files:
    fname = f.split("/")[-1]
    print(f"{'=' * 60}")
    print(f"FILE: {fname}")
    print(f"{'=' * 60}")
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        print(f"  Sheets: {wb.sheetnames}")

        for sheet_name in wb.sheetnames[:3]:  # First 3 sheets
            ws = wb[sheet_name]
            print(f"\n  --- Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) ---")

            # Get headers from row 1
            headers = []
            for cell in next(ws.iter_rows(max_row=1)):
                if cell.value:
                    headers.append((cell.column_letter, str(cell.value)[:50]))
            print(f"  Headers: {headers[:20]}")
            if len(headers) > 20:
                print(f"           ... +{len(headers) - 20} more")

            # Sample rows 2-4
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4)):
                vals = [(c.column_letter, c.value) for c in row if c.value is not None]
                print(f"  Row {i+2}: {vals[:15]}")

        wb.close()
    except Exception as e:
        print(f"  ERROR: {e}")
    print()

# Also check the CSV
print(f"{'=' * 60}")
print("FILE: FlowsheetHomeCSV.csv")
print(f"{'=' * 60}")
try:
    import csv
    with open('/Users/woleakpose/Documents/FlowsheetHomeCSV.csv', 'r') as f2:
        reader = csv.reader(f2)
        for i, row in enumerate(reader):
            if i < 5:
                print(f"  Row {i+1}: {row[:15]}")
            else:
                break
except Exception as e:
    print(f"  ERROR: {e}")
