#!/usr/bin/env python3
"""
Import hemodialysis flowsheet data from Excel workbooks into the database.

Supports these workbooks:
  - Flowsheets.xlsx (2017-2018, ~280 sessions)
  - 2023FlowSheets-ex.xlsx (2023, ~60 sessions)
  - 2024FlowSheets.xlsx (2024, ~150 sessions)
  - 2025FlowSheets.xlsx (2025, ~130 sessions)
  - FlowsheetGermantown.xlsx (when synced from OneDrive)

All use the same DaVita Home HHD Flowsheet template layout:
  Row 4/H4  = Date
  Row 6/H6  = Flow Fraction
  Row 7/L7  = Day of Week
  Row 11    = Pre Sitting BP (B=sys, C=dia, D=pulse), Symptoms
  Row 12    = Pre Standing BP (B=sys, C=dia, D=pulse), Symptoms
  Row 13    = Temp (C13), Equipment
  Row 14    = Previous Post Weight (C14), Equipment
  Row 15    = Today's Weight (C15), Equipment
  Row 16    = Dry Weight (C16)
  Row 17    = Fluid to Remove (C17), Alarm Test (V17)
  Row 19    = Filtration Fract (O19), SAK Use # (U19)
  Row 20    = Lactate (I20), K+ (K20), Access (O20), Chloramine (U20)
  Row 21    = SAK# (I21), Needle Gauge (O21), Length (Q21)
  Row 22    = Blood Flow Rate (J22), Buttonhole (O22)
  Row 24    = Drugs (row 24-28), Start Time (S24), Stop Time (U24)
  Rows 31-42 = Intradialytic readings
  Row 43    = Post sitting BP (B=sys, C=dia, D=pulse)
  Row 48-52 = Post Treatment data
  Row 49    = Post sitting BP (B49=sys, C49=dia, D49=pulse)
  Row 50    = Post standing BP (B50=sys, C50=dia, D50=pulse)
  Row 51    = Post Temp (B51), Post Weight (F51)
  Row 52    = Total Time (C52), Total Dialysate (E52 or F52), Total UF (G52 or H52)

Also imports Vomit Log sheets into vomiting_logs table.
"""

import psycopg2
import openpyxl
from datetime import datetime, date, time, timedelta
import os
import re
import traceback

# Database connection
DB_PARAMS = {
    'host': 'localhost',
    'port': 5432,
    'dbname': 'alafia',
    'user': 'alafia',
    'password': 'alafia'
}

USER_ID = 1
CONDITION_ID = 14  # ESRD

# Skip these sheet names — they aren't sessions
SKIP_SHEETS = {
    'Sheet1', 'Sheet2', 'Sheet3', 'Summary', 'Dummy', 'Template', 'template',
    'VomitLog', 'Vomit Log', 'VomitLog2023', 'VomitLog2025', 'Vomit Log 2025',
    'VomitLog2021', 'VomitLog2022',
    'Access Maintenance', 'AccessMaintenance', 'Supplies Request',
    'Maintenance Log', 'Cleaning Log', 'Supplies',
}

# All flowsheet workbooks to import
WORKBOOKS = [
    {
        'path': '/Users/woleakpose/Documents/Developer/WellnessScore/Flowsheets.xlsx',
        'label': 'Flowsheets 2017-2018',
    },
    {
        'path': '/Users/woleakpose/Documents/2023FlowSheets-ex.xlsx',
        'label': '2023 FlowSheets',
    },
    {
        'path': '/Users/woleakpose/Documents/Developer/data/2024FlowSheets.xlsx',
        'label': '2024 FlowSheets',
    },
    {
        'path': '/Users/woleakpose/Library/Group Containers/UBF8T346G9.OneDriveStandaloneSuite/OneDrive.noindex/OneDrive/2025FlowSheets.xlsx',
        'label': '2025 FlowSheets',
    },
    {
        'path': '/Users/woleakpose/Desktop/FlowsheetGermantown.xlsx',
        'label': 'FlowsheetGermantown 2019-2022',
    },
]


def cell(ws, row, col):
    """Get cell value by row number and column letter."""
    try:
        return ws[f'{col}{row}'].value
    except Exception:
        return None


def safe_int(val):
    """Convert to int, return None if not possible."""
    if val is None:
        return None
    try:
        v = int(float(val))
        return v if v != 0 else None
    except (ValueError, TypeError):
        return None


def safe_float(val):
    """Convert to float, return None if not possible."""
    if val is None:
        return None
    try:
        v = float(val)
        return v if v != 0.0 else None
    except (ValueError, TypeError):
        return None


def safe_str(val):
    """Convert to string, return None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def yes_no_bool(val):
    """Convert YES/NO/X to boolean."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ('YES', 'Y', 'TRUE', '1', 'X'):
        return True
    if s in ('NO', 'N', 'FALSE', '0'):
        return False
    return None


def parse_session_date(sheet_name, ws):
    """Parse session date from sheet name or H4 cell.
    
    Returns (date, session_number) where session_number is 1 for normal,
    2 for '.2' suffix sheets (same-day double sessions).
    """
    # Detect .2 suffix for same-day double sessions
    name = sheet_name.strip()
    session_number = 1
    if name.endswith('.2'):
        session_number = 2
        name = name[:-2]  # Strip .2 for date parsing

    # First try the H4 cell (most reliable)
    h4 = cell(ws, 4, 'H')
    if isinstance(h4, datetime):
        return h4.date(), session_number
    if isinstance(h4, date):
        return h4, session_number

    # Try parsing sheet name in various formats

    # MM-DD-YYYY format (2024/2025/Germantown workbooks)
    m = re.match(r'(\d{1,2})-(\d{1,2})-(\d{4})', name)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(1)), int(m.group(2))), session_number
        except ValueError:
            pass

    # "Mon DD-YYYY" or "Mon DD, YYYY" format (2017-2018 workbook)
    months = {
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
        'january': 1, 'february': 2, 'march': 3, 'april': 4,
        'june': 6, 'july': 7, 'august': 8, 'september': 9,
        'october': 10, 'november': 11, 'december': 12,
    }
    m = re.match(r'([A-Za-z]+)\s+(\d{1,2})[-,\s]+(\d{4})', name)
    if m:
        month_str = m.group(1).lower()
        if month_str in months:
            try:
                return date(int(m.group(3)), months[month_str], int(m.group(2))), session_number
            except ValueError:
                pass

    # "Mon DD-YY" format
    m = re.match(r'([A-Za-z]+)\s+(\d{1,2})[-,\s]+(\d{2})$', name)
    if m:
        month_str = m.group(1).lower()
        if month_str in months:
            yr = int(m.group(3))
            yr = yr + 2000 if yr < 100 else yr
            try:
                return date(yr, months[month_str], int(m.group(2))), session_number
            except ValueError:
                pass

    return None, session_number


def parse_time_value(val):
    """Parse a time value from cell."""
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        # Try HH:MM format
        m = re.match(r'(\d{1,2}):(\d{2})', val)
        if m:
            return time(int(m.group(1)), int(m.group(2)))
    return None


def extract_session_data(ws, sheet_name):
    """Extract all session data from a worksheet."""
    session_date, session_number = parse_session_date(sheet_name, ws)
    if not session_date:
        return None, None, 1

    # --- Pre-Treatment Vitals ---
    pre_sit_sys = safe_int(cell(ws, 11, 'B'))
    pre_sit_dia = safe_int(cell(ws, 11, 'C'))
    pre_sit_pulse = safe_int(cell(ws, 11, 'D'))

    pre_stand_sys = safe_int(cell(ws, 12, 'B'))
    pre_stand_dia = safe_int(cell(ws, 12, 'C'))
    pre_stand_pulse = safe_int(cell(ws, 12, 'D'))

    pre_temp = safe_float(cell(ws, 13, 'C'))
    prev_post_weight = safe_float(cell(ws, 14, 'C'))
    todays_weight = safe_float(cell(ws, 15, 'C'))
    dry_weight = safe_float(cell(ws, 16, 'C'))
    fluid_to_remove = safe_float(cell(ws, 17, 'C'))

    # Pre-treatment symptoms
    pre_sob = yes_no_bool(cell(ws, 11, 'M'))
    pre_mobility = yes_no_bool(cell(ws, 11, 'R'))
    pre_swelling = yes_no_bool(cell(ws, 12, 'M'))
    pre_digestion = yes_no_bool(cell(ws, 12, 'R'))
    pre_hosp = yes_no_bool(cell(ws, 13, 'M'))

    # Access assessment
    thrill_bruit = yes_no_bool(cell(ws, 14, 'R'))
    access_redness = yes_no_bool(cell(ws, 15, 'R'))

    # Equipment
    cartridge_lot = safe_str(cell(ws, 11, 'U'))
    sak_lot = safe_str(cell(ws, 12, 'U'))
    cycler_num = safe_str(cell(ws, 13, 'U'))
    warmer_serial = safe_str(cell(ws, 14, 'U'))
    control_panel = safe_str(cell(ws, 15, 'U'))
    alarm_test = yes_no_bool(cell(ws, 17, 'V'))

    # Flow fraction
    flow_fraction = safe_float(cell(ws, 6, 'H'))
    day_of_week = safe_str(cell(ws, 7, 'L'))

    # Dialysis Prescription
    volume_str = safe_str(cell(ws, 19, 'I'))  # "30 L"
    dialysate_vol = None
    if volume_str:
        m = re.search(r'([\d.]+)', volume_str)
        if m:
            dialysate_vol = float(m.group(1))

    filt_fract = safe_str(cell(ws, 19, 'O'))  # "48% 40%"
    sak_use = safe_int(cell(ws, 19, 'U'))

    lactate_str = safe_str(cell(ws, 20, 'I'))  # "45 mEq"
    lactate = None
    if lactate_str:
        m = re.search(r'([\d.]+)', lactate_str)
        if m:
            lactate = float(m.group(1))

    k_str = safe_str(cell(ws, 20, 'K'))  # "1.0 mEq"
    potassium = None
    if k_str:
        m = re.search(r'([\d.]+)', k_str)
        if m:
            potassium = float(m.group(1))

    access_type = safe_str(cell(ws, 20, 'O'))  # "Catheter. URJ"
    chloramine_val = cell(ws, 20, 'U')
    total_chloramine = safe_float(chloramine_val) if chloramine_val not in (None, 'NA', 'na', 'N/A') else None

    sak_number = safe_int(cell(ws, 21, 'I'))
    needle_gauge_str = safe_str(cell(ws, 21, 'O'))  # "17g"
    needle_length = safe_float(cell(ws, 21, 'Q'))
    buttonhole = yes_no_bool(cell(ws, 22, 'O'))

    bfr_str = safe_str(cell(ws, 22, 'J'))  # "350 - 450 ml/t"
    blood_flow_rate = None
    if bfr_str:
        m = re.search(r'([\d.]+)', bfr_str)
        if m:
            blood_flow_rate = float(m.group(1))

    # Drugs
    drugs = []
    for r in range(24, 29):
        drug = safe_str(cell(ws, r, 'H'))
        dose = safe_str(cell(ws, r, 'K'))
        if drug:
            drugs.append(f"{drug} ({dose})" if dose else drug)
    drugs_str = '; '.join(drugs) if drugs else None

    # Start/Stop times
    start_time_val = cell(ws, 24, 'S') or cell(ws, 24, 'T')
    stop_time_val = cell(ws, 24, 'U') or cell(ws, 24, 'V')

    start_time = parse_time_value(start_time_val)
    stop_time = parse_time_value(stop_time_val)

    actual_start = None
    actual_end = None
    if start_time:
        actual_start = datetime.combine(session_date, start_time)
    if stop_time:
        actual_end = datetime.combine(session_date, stop_time)
        # If stop time is before start time, it's the next day
        if actual_start and actual_end < actual_start:
            actual_end += timedelta(days=1)

    # Duration from row 52
    total_time_val = cell(ws, 52, 'C')
    duration_minutes = None
    t = parse_time_value(total_time_val)
    if t:
        duration_minutes = t.hour * 60 + t.minute

    # --- Post-Treatment ---
    post_sit_sys = safe_int(cell(ws, 49, 'B'))
    post_sit_dia = safe_int(cell(ws, 49, 'C'))
    post_sit_pulse = safe_int(cell(ws, 49, 'D'))

    post_stand_sys = safe_int(cell(ws, 50, 'B'))
    post_stand_dia = safe_int(cell(ws, 50, 'C'))
    post_stand_pulse = safe_int(cell(ws, 50, 'D'))

    post_temp = safe_float(cell(ws, 51, 'B'))
    post_weight = safe_float(cell(ws, 51, 'F'))

    post_sob = yes_no_bool(cell(ws, 49, 'I'))
    post_thrill = yes_no_bool(cell(ws, 49, 'L'))
    post_swelling = yes_no_bool(cell(ws, 50, 'I'))
    post_digestion = yes_no_bool(cell(ws, 51, 'I'))

    dialyzer_post = safe_str(cell(ws, 52, 'P'))

    # Machine maintenance
    purif_pak = yes_no_bool(cell(ws, 49, 'V'))
    air_filter = yes_no_bool(cell(ws, 50, 'V'))
    waste_bleach = yes_no_bool(cell(ws, 51, 'V'))

    # Total UF, dialysate
    # These can be in different columns depending on workbook
    total_dialysate = safe_float(cell(ws, 52, 'F')) or safe_float(cell(ws, 52, 'E'))
    total_uf = safe_float(cell(ws, 52, 'H')) or safe_float(cell(ws, 52, 'G'))
    total_bvp = safe_float(cell(ws, 52, 'L')) or safe_float(cell(ws, 52, 'K'))

    # Fluid removed
    fluid_removed = None
    if todays_weight and post_weight:
        fluid_removed = round((todays_weight - post_weight) * 1000, 1)  # kg to ml

    # Lab tubes
    tubes = []
    for r in [19, 20, 21]:
        t1 = safe_str(cell(ws, r, 'A'))
        t2 = safe_str(cell(ws, r, 'D'))
        if t1 and t1 not in ('Dialysate / Water Testing',):
            tubes.append(t1)
        if t2:
            tubes.append(t2)
    lab_tubes = ', '.join(tubes) if tubes else None

    # Notes (row 26)
    notes = safe_str(cell(ws, 26, 'A')) or safe_str(cell(ws, 26, 'B'))

    # RN Reviewer
    rn_reviewer = safe_str(cell(ws, 5, 'P')) or safe_str(cell(ws, 21, 'U'))

    session = {
        'user_id': USER_ID,
        'condition_id': CONDITION_ID,
        'therapy_type': 'HEMODIALYSIS',
        'therapy_name': 'Home Hemodialysis (HHD)',
        'scheduled_date': datetime.combine(session_date, time(0, 0)),
        'actual_start_time': actual_start,
        'actual_end_time': actual_end,
        'duration_minutes': duration_minutes,
        'status': 'COMPLETED',
        'facility_name': 'DaVita Home',
        'attending_physician': 'Dr. Anand Gaurang Desai',
        'dialysis_access_type': access_type,
        'pre_dialysis_weight_kg': todays_weight,
        'post_dialysis_weight_kg': post_weight,
        'fluid_removed_ml': fluid_removed,
        'blood_flow_rate': blood_flow_rate,
        'dry_weight_kg': dry_weight,
        'previous_post_weight_kg': prev_post_weight,
        'fluid_to_remove_kg': fluid_to_remove,
        'flow_fraction': flow_fraction,
        'filtration_fraction': filt_fract,
        'day_of_week': day_of_week,
        'rn_reviewer': rn_reviewer,
        'dialysate_volume_liters': dialysate_vol,
        'dialysate_lactate_meq': lactate,
        'dialysate_potassium_meq': potassium,
        'sak_number': sak_number,
        'needle_gauge': needle_gauge_str,
        'needle_length': needle_length,
        'buttonhole_technique': buttonhole,
        'pre_systolic_bp': pre_sit_sys,
        'pre_diastolic_bp': pre_sit_dia,
        'pre_heart_rate': pre_sit_pulse,
        'pre_standing_systolic_bp': pre_stand_sys,
        'pre_standing_diastolic_bp': pre_stand_dia,
        'pre_standing_heart_rate': pre_stand_pulse,
        'post_systolic_bp': post_sit_sys,
        'post_diastolic_bp': post_sit_dia,
        'post_heart_rate': post_sit_pulse,
        'post_standing_systolic_bp': post_stand_sys,
        'post_standing_diastolic_bp': post_stand_dia,
        'post_standing_heart_rate': post_stand_pulse,
        'pre_temperature': pre_temp,
        'post_temperature': post_temp,
        'cartridge_lot': cartridge_lot,
        'sak_lot': sak_lot,
        'cycler_number': cycler_num,
        'warmer_serial': warmer_serial,
        'control_panel_serial': control_panel,
        'pre_shortness_of_breath': pre_sob,
        'pre_swelling': pre_swelling,
        'pre_change_in_mobility': pre_mobility,
        'pre_digestion_problems': pre_digestion,
        'pre_hosp_er_since_last': pre_hosp,
        'access_thrill_bruit': thrill_bruit,
        'access_redness_drainage': access_redness,
        'total_dialysate_liters': total_dialysate,
        'total_uf_liters': total_uf,
        'total_blood_volume_processed': total_bvp,
        'dialyzer_appearance': dialyzer_post,
        'post_shortness_of_breath': post_sob,
        'post_swelling': post_swelling,
        'post_digestion_problems': post_digestion,
        'post_access_thrill_bruit': post_thrill,
        'purification_pak_change': purif_pak,
        'air_filter_cleaned': air_filter,
        'waste_line_bleach_disinfection': waste_bleach,
        'sak_use_number': sak_use,
        'total_chloramine_level': total_chloramine,
        'alarm_test_completed': alarm_test,
        'lab_tubes_drawn': lab_tubes,
        'drugs_administered': drugs_str,
        'clinical_notes': notes,
        'created_at': datetime.utcnow(),
        'updated_at': datetime.utcnow(),
    }

    # --- Intradialytic Readings (rows 31-42) ---
    readings = []
    reading_num = 0
    for r in range(31, 43):
        reading_time = cell(ws, r, 'A')
        sys_bp = safe_int(cell(ws, r, 'B'))
        dia_bp = safe_int(cell(ws, r, 'C'))
        pulse = safe_int(cell(ws, r, 'D'))
        dial_rate = safe_float(cell(ws, r, 'E'))
        dial_vol = safe_float(cell(ws, r, 'F'))
        uf_rate = safe_float(cell(ws, r, 'G'))
        uf_vol = safe_float(cell(ws, r, 'H'))
        bfr = safe_float(cell(ws, r, 'I'))
        art_p = safe_float(cell(ws, r, 'J'))
        ven_p = safe_float(cell(ws, r, 'K'))
        eff_p = safe_float(cell(ws, r, 'L'))
        acc_state = safe_str(cell(ws, r, 'M'))
        saline = safe_str(cell(ws, r, 'N'))
        remarks = safe_str(cell(ws, r, 'O'))
        map_val = safe_float(cell(ws, r, 'T'))

        # Only include if there's meaningful data (not just zeros)
        has_data = any([
            reading_time and isinstance(reading_time, (time, datetime)),
            sys_bp, dia_bp, pulse, dial_rate, bfr, art_p, ven_p, eff_p,
            acc_state, saline, remarks
        ])

        if has_data:
            reading_num += 1
            t = parse_time_value(reading_time)
            if t is None:
                # Use session start time offset if no time recorded
                if start_time:
                    t = time(
                        (start_time.hour + reading_num) % 24,
                        start_time.minute
                    )
                else:
                    t = time(0, 0)

            readings.append({
                'user_id': USER_ID,
                'reading_time': t,
                'reading_number': reading_num,
                'systolic_bp': sys_bp,
                'diastolic_bp': dia_bp,
                'pulse': pulse,
                'mean_arterial_pressure': map_val,
                'dialysate_rate': dial_rate,
                'dialysate_volume_remaining': dial_vol,
                'uf_rate': uf_rate,
                'uf_volume_removed': uf_vol,
                'blood_flow_rate': bfr,
                'arterial_pressure': art_p,
                'venous_pressure': ven_p,
                'effluent_pressure': eff_p,
                'access_state': acc_state,
                'saline_amount': saline,
                'remarks': remarks,
                'created_at': datetime.utcnow(),
            })

    return session, readings, session_number


def extract_vomit_log(ws):
    """Extract vomit log entries from a Vomit Log sheet."""
    entries = []
    for row in ws.iter_rows(min_row=3, values_only=False):
        vals = {c.column_letter: c.value for c in row if c.value is not None}
        if 'A' not in vals:
            continue

        vomit_date = vals.get('A')
        if not isinstance(vomit_date, (datetime, date)):
            continue

        vomit_time = vals.get('B')
        if isinstance(vomit_date, datetime):
            vomit_date = vomit_date.date()

        pre_weight = safe_float(vals.get('C'))
        post_weight = safe_float(vals.get('D'))

        # Auto-detect column order: some workbooks have E=delta F=description,
        # others (Germantown) have E=description F=delta
        e_val = vals.get('E')
        f_val = vals.get('F')
        if isinstance(e_val, (int, float)) and not isinstance(e_val, bool):
            delta = safe_float(e_val)
            visual = safe_str(f_val)
        else:
            visual = safe_str(e_val)
            delta = safe_float(f_val)

        last_meal_time = vals.get('G')
        last_dialysis_date = vals.get('H')
        last_dialysis_time = vals.get('I')

        # Calculate volume from weight delta (kg to ml)
        volume_ml = None
        if delta:
            volume_ml = abs(delta) * 1000

        t = parse_time_value(vomit_time) if vomit_time else time(0, 0)
        occurred_at = datetime.combine(vomit_date, t) if t else datetime.combine(vomit_date, time(0, 0))

        entries.append({
            'user_id': USER_ID,
            'occurred_at': occurred_at,
            'volume_ml': volume_ml,
            'description': visual,
            'pre_weight_kg': pre_weight,
            'post_weight_kg': post_weight,
            'notes': f"Last meal: {last_meal_time}, Last dialysis: {last_dialysis_date} {last_dialysis_time}" if last_meal_time else None,
            'created_at': datetime.utcnow(),
            'updated_at': datetime.utcnow(),
        })

    return entries


def import_workbook(conn, wb_info):
    """Import all sessions from one workbook."""
    path = wb_info['path']
    label = wb_info['label']

    if not os.path.exists(path):
        print(f"  SKIP: File not found: {path}")
        return 0, 0, 0

    # Check if it's a valid xlsx
    import subprocess
    result = subprocess.run(['file', path], capture_output=True, text=True)
    if 'Microsoft Excel' not in result.stdout and 'Zip' not in result.stdout and 'zip' not in result.stdout.lower():
        print(f"  SKIP: Not a valid xlsx file (OneDrive stub?): {path}")
        return 0, 0, 0

    print(f"\n{'='*60}")
    print(f"IMPORTING: {label}")
    print(f"  Path: {path}")

    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    except Exception as e:
        print(f"  ERROR loading workbook: {e}")
        return 0, 0, 0

    sheets = wb.sheetnames
    session_sheets = [s for s in sheets if s not in SKIP_SHEETS]
    vomit_sheets = [s for s in sheets if 'vomit' in s.lower() or 'Vomit' in s]

    print(f"  Total sheets: {len(sheets)}")
    print(f"  Session sheets: {len(session_sheets)}")
    print(f"  Vomit log sheets: {len(vomit_sheets)}")

    cur = conn.cursor()

    # Get existing session dates to avoid duplicates
    # Use full scheduled_date (datetime) to allow multiple sessions per day
    cur.execute("""
        SELECT scheduled_date FROM therapy_sessions
        WHERE user_id = %s AND therapy_type = 'HEMODIALYSIS'
    """, (USER_ID,))
    existing_timestamps = {row[0] for row in cur.fetchall()}
    print(f"  Existing sessions in DB: {len(existing_timestamps)}")

    sessions_imported = 0
    readings_imported = 0
    vomits_imported = 0

    # Import sessions
    for sheet_name in session_sheets:
        try:
            ws = wb[sheet_name]
            session, readings, session_number = extract_session_data(ws, sheet_name)

            if session is None:
                continue

            # For same-day double sessions (.2 suffix), offset scheduled_date by 12 hours
            if session_number == 2:
                session['scheduled_date'] = datetime.combine(
                    session['scheduled_date'].date(), time(12, 0)
                )

            if session['scheduled_date'] in existing_timestamps:
                continue  # Skip duplicate

            # Insert session
            cols = list(session.keys())
            vals = [session[c] for c in cols]
            placeholders = ', '.join(['%s'] * len(cols))
            col_names = ', '.join(cols)

            cur.execute(
                f"INSERT INTO therapy_sessions ({col_names}) VALUES ({placeholders}) RETURNING id",
                vals
            )
            session_id = cur.fetchone()[0]
            existing_timestamps.add(session['scheduled_date'])
            sessions_imported += 1

            # Insert intradialytic readings
            for reading in readings:
                reading['session_id'] = session_id
                rcols = list(reading.keys())
                rvals = [reading[c] for c in rcols]
                rplaceholders = ', '.join(['%s'] * len(rcols))
                rcol_names = ', '.join(rcols)
                cur.execute(
                    f"INSERT INTO intradialytic_readings ({rcol_names}) VALUES ({rplaceholders})",
                    rvals
                )
                readings_imported += 1

        except Exception as e:
            print(f"  ERROR on sheet '{sheet_name}': {e}")
            traceback.print_exc()
            conn.rollback()
            # Re-get cursor after rollback
            cur = conn.cursor()
            # Re-fetch existing timestamps
            cur.execute("""
                SELECT scheduled_date FROM therapy_sessions
                WHERE user_id = %s AND therapy_type = 'HEMODIALYSIS'
            """, (USER_ID,))
            existing_timestamps = {row[0] for row in cur.fetchall()}
            continue

    conn.commit()

    # Import vomit logs
    if vomit_sheets:
        # Get existing vomit dates to avoid duplicates
        cur.execute("""
            SELECT log_date, log_time FROM vomiting_logs WHERE user_id = %s
        """, (USER_ID,))
        existing_vomits = {(row[0], row[1]) for row in cur.fetchall()}

        for vs in vomit_sheets:
            try:
                ws = wb[vs]
                entries = extract_vomit_log(ws)
                for entry in entries:
                    log_dt = entry['occurred_at']
                    dedup_key = (log_dt.date(), log_dt.time())
                    if dedup_key in existing_vomits:
                        continue
                    # Map to actual vomiting_logs columns
                    insert_cols = ['user_id', 'log_date', 'log_time', 'volume', 'notes', 'created_at']
                    insert_vals = [
                        entry['user_id'],
                        log_dt.date(),
                        log_dt.time(),
                        entry.get('volume_ml'),
                        entry.get('notes'),
                        entry['created_at'],
                    ]
                    placeholders = ', '.join(['%s'] * len(insert_cols))
                    col_names = ', '.join(insert_cols)
                    cur.execute(
                        f"INSERT INTO vomiting_logs ({col_names}) VALUES ({placeholders})",
                        insert_vals
                    )
                    existing_vomits.add(dedup_key)
                    vomits_imported += 1
            except Exception as e:
                print(f"  ERROR on vomit sheet '{vs}': {e}")
                traceback.print_exc()
                conn.rollback()
                cur = conn.cursor()
                continue

        conn.commit()

    wb.close()
    print(f"  Imported: {sessions_imported} sessions, {readings_imported} readings, {vomits_imported} vomit entries")
    return sessions_imported, readings_imported, vomits_imported


def main():
    print("=" * 60)
    print("ALAFIA Hemodialysis Flowsheet Import")
    print("=" * 60)

    conn = psycopg2.connect(**DB_PARAMS)

    # Check current state
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM therapy_sessions WHERE user_id = %s", (USER_ID,))
    existing_sessions = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM intradialytic_readings WHERE user_id = %s", (USER_ID,))
    existing_readings = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM vomiting_logs WHERE user_id = %s", (USER_ID,))
    existing_vomits = cur.fetchone()[0]
    print(f"\nCurrent DB state:")
    print(f"  Therapy sessions: {existing_sessions}")
    print(f"  Intradialytic readings: {existing_readings}")
    print(f"  Vomiting logs: {existing_vomits}")

    # First, check vomiting_logs table columns
    cur.execute("""
        SELECT column_name FROM information_schema.columns
        WHERE table_name = 'vomiting_logs' ORDER BY ordinal_position
    """)
    vomit_cols = [row[0] for row in cur.fetchall()]
    print(f"\n  Vomiting log columns: {vomit_cols}")

    total_sessions = 0
    total_readings = 0
    total_vomits = 0

    for wb_info in WORKBOOKS:
        s, r, v = import_workbook(conn, wb_info)
        total_sessions += s
        total_readings += r
        total_vomits += v

    # Final state
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM therapy_sessions WHERE user_id = %s", (USER_ID,))
    final_sessions = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM intradialytic_readings WHERE user_id = %s", (USER_ID,))
    final_readings = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM vomiting_logs WHERE user_id = %s", (USER_ID,))
    final_vomits = cur.fetchone()[0]

    cur.execute("""
        SELECT MIN(DATE(scheduled_date)), MAX(DATE(scheduled_date))
        FROM therapy_sessions WHERE user_id = %s AND therapy_type = 'HEMODIALYSIS'
    """, (USER_ID,))
    date_range = cur.fetchone()

    conn.close()

    print(f"\n{'='*60}")
    print(f"IMPORT COMPLETE")
    print(f"{'='*60}")
    print(f"  New sessions imported: {total_sessions}")
    print(f"  New readings imported: {total_readings}")
    print(f"  New vomit entries imported: {total_vomits}")
    print(f"\n  Total therapy sessions: {final_sessions}")
    print(f"  Total intradialytic readings: {final_readings}")
    print(f"  Total vomiting logs: {final_vomits}")
    if date_range[0]:
        print(f"  Date range: {date_range[0]} to {date_range[1]}")


if __name__ == '__main__':
    main()
