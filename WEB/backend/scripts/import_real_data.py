#!/usr/bin/env python3
"""
Import real patient data from Records.xlsx into ALAFIA database.

Data sources:
- Lab sheet: 168 lab tests × 296 dates (2016-2024), pivoted format
- FoodLog sheets: FoodLog (2021), FoodLog2022, FoodLog2023, FoodLog2024
- Poop Log: 1,188 bowel movement entries (2022-2024)

Connects directly to PostgreSQL for batch inserts (much faster than API calls).
"""

import sys
import os
from datetime import datetime, date, time, timedelta
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ── Configuration ──────────────────────────────────────────────────────────
EXCEL_PATH = os.path.expanduser(
    "~/Library/Mobile Documents/com~apple~CloudDocs/Records.xlsx"
)
DB_CONFIG = {
    "host": "localhost",
    "port": 5432,
    "dbname": "alafia",
    "user": "alafia",
    "password": "alafia",
}
USER_ID = 1  # demo@alafia.app

# Lab test → category mapping
LAB_CATEGORIES = {
    # Kidney function
    "BUN": "Kidney", "Creatinine": "Kidney", "eGFR": "Kidney",
    "BUN/Creatinine Ratio": "Kidney", "Urea Nitrogen": "Kidney",
    "Cystatin C": "Kidney", "Microalbumin": "Kidney",
    "Protein, Total, Urine": "Kidney", "Creatinine, Urine": "Kidney",
    # CBC / Hematology
    "WBC": "Hematology", "RBC": "Hematology", "Hemoglobin": "Hematology",
    "Hematocrit": "Hematology", "MCV": "Hematology", "MCH": "Hematology",
    "MCHC": "Hematology", "RDW": "Hematology", "Platelet Count": "Hematology",
    "MPV": "Hematology", "Neutrophils": "Hematology", "Lymphocytes": "Hematology",
    "Monocytes": "Hematology", "Eosinophils": "Hematology", "Basophils": "Hematology",
    "Neutrophils (Absolute)": "Hematology", "Lymphs (Absolute)": "Hematology",
    "Monocytes(Absolute)": "Hematology", "Eos (Absolute)": "Hematology",
    "Baso (Absolute)": "Hematology", "Immature Granulocytes": "Hematology",
    "Immature Grans (Abs)": "Hematology", "NRBC": "Hematology",
    # Hemolysis markers (critical for G6PD story)
    "Reticulocyte Count": "Hemolysis", "Reticulocyte, Absolute": "Hemolysis",
    "LDH": "Hemolysis", "Haptoglobin": "Hemolysis", "Bilirubin, Total": "Hemolysis",
    "Bilirubin, Direct": "Hemolysis", "Bilirubin, Indirect": "Hemolysis",
    "G6PD RBC": "Hemolysis",
    # Iron studies
    "Iron": "Iron Studies", "Iron Bind.Cap.(TIBC)": "Iron Studies",
    "UIBC": "Iron Studies", "Iron Saturation": "Iron Studies",
    "Ferritin": "Iron Studies", "Transferrin": "Iron Studies",
    # Metabolic panel
    "Sodium": "Metabolic", "Potassium": "Metabolic", "Chloride": "Metabolic",
    "Carbon Dioxide, Total": "Metabolic", "CO2": "Metabolic",
    "Calcium": "Metabolic", "Phosphorus": "Metabolic",
    "Glucose": "Metabolic", "Magnesium": "Metabolic", "Uric Acid": "Metabolic",
    "Anion Gap": "Metabolic", "Osmolality": "Metabolic",
    # Liver
    "AST (SGOT)": "Liver", "ALT (SGPT)": "Liver", "Alkaline Phosphatase": "Liver",
    "Protein, Total": "Liver", "Albumin": "Liver", "Globulin": "Liver",
    "A/G Ratio": "Liver", "GGT": "Liver",
    # Lipid panel
    "Cholesterol, Total": "Lipids", "Triglycerides": "Lipids",
    "HDL Cholesterol": "Lipids", "VLDL Cholesterol Cal": "Lipids",
    "LDL Cholesterol Calc": "Lipids", "LDL/HDL Ratio": "Lipids",
    "Chol/HDLC Ratio": "Lipids",
    # Thyroid
    "TSH": "Thyroid", "T4, Free(Direct)": "Thyroid", "T3, Free": "Thyroid",
    "Thyroxine (T4)": "Thyroid", "T3 Uptake": "Thyroid",
    "Free Thyroxine Index": "Thyroid",
    # Bone/Mineral
    "PTH, Intact": "Bone/Mineral", "Vitamin D, 25-Hydroxy": "Bone/Mineral",
    # Coagulation
    "PT": "Coagulation", "INR": "Coagulation", "PTT": "Coagulation",
    # Diabetes
    "Hemoglobin A1c": "Diabetes", "eAG": "Diabetes",
    # Infectious disease
    "Hepatitis B Surface Ag": "Infectious", "Hepatitis B Surface Ab": "Infectious",
    "Hepatitis C Antibody": "Infectious", "HIV 1/O/2 Abs": "Infectious",
    # Cardiac
    "Troponin I": "Cardiac", "BNP": "Cardiac", "CK": "Cardiac",
    # Inflammation
    "C-Reactive Protein": "Inflammation", "ESR": "Inflammation",
    "Sed Rate": "Inflammation",
}


def connect_db():
    """Connect to PostgreSQL."""
    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = False
    return conn


def clear_existing_data(conn, user_id):
    """Remove existing demo data to avoid duplicates."""
    cur = conn.cursor()
    for table in ["lab_results", "nutrition_logs", "bowel_movements"]:
        cur.execute(f"DELETE FROM {table} WHERE user_id = %s", (user_id,))
        print(f"  Cleared {cur.rowcount} rows from {table}")
    conn.commit()
    cur.close()


def parse_float(val):
    """Try to parse a value as float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip().replace(",", "")
        # Handle special markers
        if val in ("", "-", "N/A", "#N/A", "---", "****", "See Below",
                    "See Note", "TNP", "QNS", "Cancel", "Cancelled"):
            return None
        # Handle < and > prefixed values
        if val.startswith("<") or val.startswith(">"):
            val = val[1:].strip()
        try:
            return float(val)
        except (ValueError, TypeError):
            return None
    return None


def import_labs(conn, wb, user_id):
    """
    Import lab results from the 'Lab' sheet.
    
    Sheet structure (pivoted):
    - Row 1: header labels
    - Row 2: dates (columns F onward)
    - Row 3: possibly empty
    - Row 4: facility names (columns F onward)
    - Rows 5+: test values
    - Column A: test name
    - Column B: unit (sometimes)
    - Column C: (varies)
    - Column D: reference range low
    - Column E: reference range high
    """
    print("\n" + "=" * 60)
    print("IMPORTING LAB RESULTS")
    print("=" * 60)

    ws = wb["Lab"]
    max_row = ws.max_row
    max_col = ws.max_column

    # Step 1: Find date columns by scanning row 2
    date_cols = {}  # col_idx -> date
    facility_cols = {}  # col_idx -> facility name
    
    for col in range(1, max_col + 1):
        cell_val = ws.cell(row=2, column=col).value
        if isinstance(cell_val, datetime):
            date_cols[col] = cell_val.date()
        elif isinstance(cell_val, date):
            date_cols[col] = cell_val

    # Step 2: Get facility names from row 4
    for col in date_cols:
        fac = ws.cell(row=4, column=col).value
        if fac and isinstance(fac, str) and fac.strip():
            facility_cols[col] = fac.strip()

    print(f"  Found {len(date_cols)} date columns")
    if date_cols:
        dates_sorted = sorted(date_cols.values())
        print(f"  Date range: {dates_sorted[0]} to {dates_sorted[-1]}")
    print(f"  Found {len(facility_cols)} facility entries")

    # Step 3: Extract test rows (starting from row 5)
    records = []
    test_count = 0
    
    for row in range(5, max_row + 1):
        test_name = ws.cell(row=row, column=1).value  # Col A
        if not test_name or not isinstance(test_name, str) or not test_name.strip():
            continue
        
        test_name = test_name.strip()
        test_count += 1
        
        # Get unit from column B
        unit_raw = ws.cell(row=row, column=2).value
        unit = str(unit_raw).strip() if unit_raw else None
        # Some "units" are just flags like "L" or "H" - skip those
        if unit and unit.upper() in ("L", "H", "LL", "HH"):
            unit = None
        
        # Get reference ranges from columns D and E
        ref_low = parse_float(ws.cell(row=row, column=4).value)
        ref_high = parse_float(ws.cell(row=row, column=5).value)
        
        # Get category
        category = LAB_CATEGORIES.get(test_name)
        # Try partial matching if exact not found
        if not category:
            for known_test, cat in LAB_CATEGORIES.items():
                if known_test.lower() in test_name.lower() or test_name.lower() in known_test.lower():
                    category = cat
                    break
        
        # Iterate over date columns
        for col, test_date in date_cols.items():
            cell_val = ws.cell(row=row, column=col).value
            if cell_val is None:
                continue
            
            # Try to parse as numeric
            value = parse_float(cell_val)
            value_string = None
            
            if value is None and cell_val is not None:
                # Store as string value if not parseable
                s = str(cell_val).strip()
                if s and s not in ("", "-", "N/A", "#N/A", "---"):
                    value_string = s
                else:
                    continue  # Skip empty/null markers
            
            # Determine if abnormal
            is_abnormal = None
            if value is not None and ref_low is not None and ref_high is not None:
                is_abnormal = value < ref_low or value > ref_high
            
            facility = facility_cols.get(col)
            
            records.append((
                user_id,
                test_date,
                test_name,
                None,  # loinc_code
                category,
                value,
                value_string,
                unit,
                ref_low,
                ref_high,
                is_abnormal,
                "final",  # status
                None,  # ordering_provider
                facility,  # performing_lab
                None,  # notes
                datetime.now(),  # created_at
            ))

    print(f"  Processed {test_count} unique tests")
    print(f"  Generated {len(records)} lab result records")

    # Step 4: Batch insert
    if records:
        cur = conn.cursor()
        insert_sql = """
            INSERT INTO lab_results (
                user_id, test_date, test_name, loinc_code, category,
                value, value_string, unit, reference_range_low, reference_range_high,
                is_abnormal, status, ordering_provider, performing_lab, notes,
                created_at
            ) VALUES %s
        """
        execute_values(cur, insert_sql, records, page_size=500)
        conn.commit()
        cur.close()
        print(f"  ✓ Inserted {len(records)} lab results")
    
    return len(records)


def classify_meal_type(t):
    """Classify meal type based on time of day."""
    if t is None:
        return "snack"
    
    if isinstance(t, datetime):
        hour = t.hour
    elif isinstance(t, time):
        hour = t.hour
    else:
        return "snack"
    
    if 5 <= hour < 11:
        return "breakfast"
    elif 11 <= hour < 15:
        return "lunch"
    elif 15 <= hour < 17:
        return "snack"
    elif 17 <= hour < 22:
        return "dinner"
    else:
        return "snack"  # late night


def import_food_logs(conn, wb, user_id):
    """
    Import food log entries from multiple sheets.
    
    Sheets and their column mappings:
    - FoodLog (2021): H=Date, I=Start, J=End, L=Description, M=Qty, N=PreWt, O=PostWt, P=State
    - FoodLog2022: A=Date, B=Start, C=End, E=Desc, F=Qty, G=PreWt, H=PostWt, I=State
    - FoodLog2023: A=Date, B=Start, C=End, E=Desc, F=Qty, G=PreWt, H=PostWt, I=State, J=Med
    - FoodLog2024: A=Date, B=Start, C=End, E=Desc, F=Qty, G=PreWt, H=PostWt, N=State, O=Med
    """
    print("\n" + "=" * 60)
    print("IMPORTING FOOD LOGS")
    print("=" * 60)

    # Define column mappings for each sheet (1-indexed column numbers)
    sheet_configs = {
        "FoodLog": {
            "date_col": 8,   # H
            "start_col": 9,  # I
            "end_col": 10,   # J
            "desc_col": 12,  # L
            "qty_col": 13,   # M
            "pre_wt_col": 14, # N
            "post_wt_col": 15, # O
            "state_col": 16,  # P
            "med_col": None,
            "start_row": 3,  # Row 2 is header-ish
        },
        "FoodLog2022": {
            "date_col": 1,   # A
            "start_col": 2,  # B
            "end_col": 3,    # C
            "desc_col": 5,   # E
            "qty_col": 6,    # F
            "pre_wt_col": 7, # G
            "post_wt_col": 8, # H
            "state_col": 9,  # I
            "med_col": None,
            "start_row": 3,  # Row 2 is header
        },
        "FoodLog2023": {
            "date_col": 1,   # A
            "start_col": 2,  # B
            "end_col": 3,    # C
            "desc_col": 5,   # E
            "qty_col": 6,    # F
            "pre_wt_col": 7, # G
            "post_wt_col": 8, # H
            "state_col": 9,  # I
            "med_col": 10,   # J
            "start_row": 2,  # Row 1 is header
        },
        "FoodLog2024": {
            "date_col": 1,   # A
            "start_col": 2,  # B
            "end_col": 3,    # C
            "desc_col": 5,   # E
            "qty_col": 6,    # F
            "pre_wt_col": 7, # G
            "post_wt_col": 8, # H
            "state_col": 14, # N
            "med_col": 15,   # O
            "start_row": 2,  # Row 1 is header
        },
    }

    total_records = 0
    all_records = []

    for sheet_name, cfg in sheet_configs.items():
        if sheet_name not in wb.sheetnames:
            print(f"  Sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]
        max_row = ws.max_row
        sheet_count = 0

        for row in range(cfg["start_row"], max_row + 1):
            # Get date
            date_val = ws.cell(row=row, column=cfg["date_col"]).value
            if date_val is None:
                continue
            if isinstance(date_val, datetime):
                log_date = date_val.date()
            elif isinstance(date_val, date):
                log_date = date_val
            else:
                continue  # Skip non-date rows

            # Get description (food_name) - required
            desc = ws.cell(row=row, column=cfg["desc_col"]).value
            if not desc or not isinstance(desc, str) or not desc.strip():
                continue
            food_name = desc.strip()
            if len(food_name) > 500:
                food_name = food_name[:497] + "..."

            # Get start time for meal classification
            start_time = ws.cell(row=row, column=cfg["start_col"]).value
            meal_type = classify_meal_type(start_time)

            # Get quantity as serving_size
            qty_val = ws.cell(row=row, column=cfg["qty_col"]).value
            serving_size = None
            if qty_val is not None:
                serving_size = str(qty_val).strip() if qty_val else None

            # Build notes from pre/post weight, state, medication, time
            notes_parts = []
            
            # Time info
            if start_time:
                if isinstance(start_time, (time, datetime)):
                    t = start_time if isinstance(start_time, time) else start_time.time()
                    notes_parts.append(f"Time: {t.strftime('%H:%M')}")
            
            end_time = ws.cell(row=row, column=cfg["end_col"]).value
            if end_time:
                if isinstance(end_time, (time, datetime)):
                    t = end_time if isinstance(end_time, time) else end_time.time()
                    notes_parts.append(f"End: {t.strftime('%H:%M')}")
            
            pre_wt = ws.cell(row=row, column=cfg["pre_wt_col"]).value
            post_wt = ws.cell(row=row, column=cfg["post_wt_col"]).value
            if pre_wt is not None and isinstance(pre_wt, (int, float)):
                notes_parts.append(f"Pre-weight: {pre_wt}kg")
            if post_wt is not None and isinstance(post_wt, (int, float)):
                notes_parts.append(f"Post-weight: {post_wt}kg")

            state_val = ws.cell(row=row, column=cfg["state_col"]).value
            if state_val and isinstance(state_val, str) and state_val.strip():
                notes_parts.append(f"State: {state_val.strip()}")

            if cfg["med_col"]:
                med_val = ws.cell(row=row, column=cfg["med_col"]).value
                if med_val and isinstance(med_val, str) and med_val.strip():
                    notes_parts.append(f"Medication: {med_val.strip()}")

            notes = " | ".join(notes_parts) if notes_parts else None

            all_records.append((
                user_id,
                log_date,
                meal_type,
                food_name,
                serving_size,
                None,  # calories
                None,  # protein_g
                None,  # carbs_g
                None,  # fat_g
                notes,
                datetime.now(),  # created_at
            ))
            sheet_count += 1

        print(f"  {sheet_name}: {sheet_count} entries")
        total_records += sheet_count

    # Batch insert
    if all_records:
        cur = conn.cursor()
        insert_sql = """
            INSERT INTO nutrition_logs (
                user_id, log_date, meal_type, food_name, serving_size,
                calories, protein_g, carbs_g, fat_g, notes,
                created_at
            ) VALUES %s
        """
        execute_values(cur, insert_sql, all_records, page_size=500)
        conn.commit()
        cur.close()
        print(f"  ✓ Inserted {len(all_records)} nutrition log entries")

    return len(all_records)


def import_poop_log(conn, wb, user_id):
    """
    Import bowel movement data from 'Poop Log' sheet.
    
    Columns: A=Date, B=Time, C=Last Meal, D=Blood?, E=Notes
    """
    print("\n" + "=" * 60)
    print("IMPORTING BOWEL MOVEMENT LOG")
    print("=" * 60)

    if "Poop Log" not in wb.sheetnames:
        print("  Sheet 'Poop Log' not found, skipping")
        return 0

    ws = wb["Poop Log"]
    max_row = ws.max_row
    records = []

    for row in range(2, max_row + 1):  # Row 1 is header
        # Date (required)
        date_val = ws.cell(row=row, column=1).value  # A
        if date_val is None:
            continue
        if isinstance(date_val, datetime):
            log_date = date_val.date()
        elif isinstance(date_val, date):
            log_date = date_val
        else:
            continue

        # Time
        time_val = ws.cell(row=row, column=2).value  # B
        log_time = None
        if isinstance(time_val, time):
            log_time = time_val
        elif isinstance(time_val, datetime):
            log_time = time_val.time()

        # Last meal time
        last_meal = ws.cell(row=row, column=3).value  # C
        last_meal_str = None
        if isinstance(last_meal, time):
            last_meal_str = last_meal.strftime("%H:%M")
        elif isinstance(last_meal, datetime):
            last_meal_str = last_meal.strftime("%H:%M")
        elif last_meal and isinstance(last_meal, str):
            last_meal_str = last_meal.strip()

        # Blood present
        blood_raw = ws.cell(row=row, column=4).value  # D
        blood_present = None
        if blood_raw and isinstance(blood_raw, str):
            blood_raw = blood_raw.strip().upper()
            if blood_raw in ("YES", "Y", "TRUE"):
                blood_present = True
            elif blood_raw in ("NO", "N", "FALSE"):
                blood_present = False

        # Notes
        notes_val = ws.cell(row=row, column=5).value  # E
        notes_parts = []
        if notes_val and isinstance(notes_val, str) and notes_val.strip():
            notes_parts.append(notes_val.strip())
        if last_meal_str:
            notes_parts.append(f"Last meal: {last_meal_str}")
        notes = " | ".join(notes_parts) if notes_parts else None

        records.append((
            user_id,
            log_date,
            log_time,
            None,  # bristol_scale
            None,  # color
            None,  # consistency
            None,  # size
            blood_present,
            None,  # mucus_present
            None,  # undigested_food
            None,  # pain_level
            None,  # straining
            None,  # urgency
            None,  # duration_minutes
            None,  # location
            None,  # associated_symptoms
            notes,
            datetime.now(),  # created_at
        ))

    print(f"  Found {len(records)} bowel movement entries")

    # Batch insert
    if records:
        cur = conn.cursor()
        insert_sql = """
            INSERT INTO bowel_movements (
                user_id, log_date, log_time, bristol_scale, color,
                consistency, size, blood_present, mucus_present,
                undigested_food, pain_level, straining, urgency,
                duration_minutes, location, associated_symptoms, notes,
                created_at
            ) VALUES %s
        """
        execute_values(cur, insert_sql, records, page_size=500)
        conn.commit()
        cur.close()
        print(f"  ✓ Inserted {len(records)} bowel movement entries")

    return len(records)


def print_summary(conn, user_id):
    """Print final data counts."""
    cur = conn.cursor()
    print("\n" + "=" * 60)
    print("FINAL DATABASE SUMMARY")
    print("=" * 60)
    
    # Lab results
    cur.execute("SELECT COUNT(*) FROM lab_results WHERE user_id = %s", (user_id,))
    lab_count = cur.fetchone()[0]
    
    cur.execute("""
        SELECT MIN(test_date), MAX(test_date), COUNT(DISTINCT test_name)
        FROM lab_results WHERE user_id = %s
    """, (user_id,))
    lab_row = cur.fetchone()
    print(f"\n  Lab Results: {lab_count} records")
    if lab_row[0]:
        print(f"    Date range: {lab_row[0]} to {lab_row[1]}")
        print(f"    Unique tests: {lab_row[2]}")
    
    # Show top categories
    cur.execute("""
        SELECT category, COUNT(*) FROM lab_results 
        WHERE user_id = %s AND category IS NOT NULL
        GROUP BY category ORDER BY COUNT(*) DESC LIMIT 10
    """, (user_id,))
    cats = cur.fetchall()
    if cats:
        print("    Top categories:")
        for cat, cnt in cats:
            print(f"      {cat}: {cnt}")
    
    # Nutrition
    cur.execute("SELECT COUNT(*) FROM nutrition_logs WHERE user_id = %s", (user_id,))
    nutr_count = cur.fetchone()[0]
    
    cur.execute("""
        SELECT MIN(log_date), MAX(log_date)
        FROM nutrition_logs WHERE user_id = %s
    """, (user_id,))
    nutr_row = cur.fetchone()
    print(f"\n  Nutrition Logs: {nutr_count} records")
    if nutr_row[0]:
        print(f"    Date range: {nutr_row[0]} to {nutr_row[1]}")
    
    cur.execute("""
        SELECT meal_type, COUNT(*) FROM nutrition_logs 
        WHERE user_id = %s GROUP BY meal_type ORDER BY COUNT(*) DESC
    """, (user_id,))
    meals = cur.fetchall()
    if meals:
        print("    By meal type:")
        for mt, cnt in meals:
            print(f"      {mt}: {cnt}")
    
    # Bowel movements
    cur.execute("SELECT COUNT(*) FROM bowel_movements WHERE user_id = %s", (user_id,))
    bm_count = cur.fetchone()[0]
    
    cur.execute("""
        SELECT MIN(log_date), MAX(log_date)
        FROM bowel_movements WHERE user_id = %s
    """, (user_id,))
    bm_row = cur.fetchone()
    print(f"\n  Bowel Movements: {bm_count} records")
    if bm_row[0]:
        print(f"    Date range: {bm_row[0]} to {bm_row[1]}")
    
    cur.execute("""
        SELECT blood_present, COUNT(*) FROM bowel_movements
        WHERE user_id = %s AND blood_present IS NOT NULL
        GROUP BY blood_present
    """, (user_id,))
    blood = cur.fetchall()
    if blood:
        print("    Blood presence:")
        for bp, cnt in blood:
            print(f"      {'Yes' if bp else 'No'}: {cnt}")
    
    # Grand total
    total = lab_count + nutr_count + bm_count
    print(f"\n  TOTAL RECORDS: {total}")
    print("=" * 60)
    
    cur.close()


def main():
    print("=" * 60)
    print("ALAFIA — Real Patient Data Import")
    print(f"Source: {EXCEL_PATH}")
    print("=" * 60)

    # Verify Excel file exists
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)

    # Load workbook (NOT read-only — read_only=True is very slow for random access)
    print("\nLoading Excel workbook (this may take a moment)...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # Connect to database
    print("\nConnecting to database...")
    conn = connect_db()
    print("  Connected successfully")

    # Clear existing data
    print("\nClearing existing demo data...")
    clear_existing_data(conn, USER_ID)

    try:
        # Import each data type
        lab_count = import_labs(conn, wb, USER_ID)
        food_count = import_food_logs(conn, wb, USER_ID)
        poop_count = import_poop_log(conn, wb, USER_ID)

        # Print final summary
        print_summary(conn, USER_ID)

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        wb.close()
        conn.close()
        print("\nDone.")


if __name__ == "__main__":
    main()
