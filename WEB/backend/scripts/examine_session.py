#!/usr/bin/env python3
"""Examine the detailed structure of a single dialysis session flowsheet."""

import openpyxl

# Look at a 2024 session (most recent format)
wb = openpyxl.load_workbook(
    '/Users/woleakpose/Documents/Developer/data/2024FlowSheets.xlsx',
    read_only=True, data_only=True
)

session_sheet = '01-02-2024'
ws = wb[session_sheet]

print(f"=== Session: {session_sheet} ({ws.max_row} rows x {ws.max_column} cols) ===\n")

print("ALL CELL DATA:")
for i, row in enumerate(ws.iter_rows(max_row=ws.max_row)):
    vals = [(c.column_letter, c.value) for c in row if c.value is not None]
    if vals:
        print(f"  Row {i+1}: {vals}")

wb.close()

# Also look at a 2018 session for comparison
print("\n\n")
wb2 = openpyxl.load_workbook(
    '/Users/woleakpose/Documents/Developer/WellnessScore/Flowsheets.xlsx',
    read_only=True, data_only=True
)

session_sheet2 = 'Nov 14-2018'
ws2 = wb2[session_sheet2]

print(f"=== Session: {session_sheet2} ({ws2.max_row} rows x {ws2.max_column} cols) ===\n")

print("ALL CELL DATA:")
for i, row in enumerate(ws2.iter_rows(max_row=ws2.max_row)):
    vals = [(c.column_letter, c.value) for c in row if c.value is not None]
    if vals:
        print(f"  Row {i+1}: {vals}")

wb2.close()
