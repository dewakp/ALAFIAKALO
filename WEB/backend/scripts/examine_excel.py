#!/usr/bin/env python3
"""Examine Records.xlsx structure for import planning."""
import openpyxl
from datetime import datetime

FILE = '/Users/woleakpose/Library/Mobile Documents/com~apple~CloudDocs/Records.xlsx'
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)

# ── Lab Sheet ────────────────────────────
ws = wb['Lab']
print("=" * 60)
print("LAB SHEET ANALYSIS")
print("=" * 60)

# Get all test names from column A
print("\nAll test names (column A):")
tests = []
for row in ws.iter_rows(min_row=5, max_col=1, values_only=True):
    if row[0]:
        tests.append(row[0])
        print(f"  {row[0]}")
print(f"\nTotal unique tests: {len(tests)}")

# Get reference data from columns B-G for first 20 tests
print("\nReference data (cols B-G) for first 20 tests:")
for row in ws.iter_rows(min_row=5, max_row=25, max_col=7, values_only=False):
    name = row[0].value
    if name:
        extras = [(c.column_letter, c.value) for c in row[1:] if c.value is not None]
        if extras:
            print(f"  {name}: {extras}")

# Count dates in row 2
print("\nDate columns (row 2):")
dates = []
for cell in list(ws.iter_rows(min_row=2, max_row=2))[0]:
    if cell.value and isinstance(cell.value, datetime):
        dates.append(cell.value)
print(f"  Total date columns: {len(dates)}")
if dates:
    print(f"  First date: {dates[0].strftime('%Y-%m-%d')}")
    print(f"  Last date: {dates[-1].strftime('%Y-%m-%d')}")
    print(f"  Date range: {(dates[-1] - dates[0]).days} days")

# Facility names in row 4
print("\nFacilities (row 4):")
facilities = set()
for cell in list(ws.iter_rows(min_row=4, max_row=4))[0]:
    if cell.value and isinstance(cell.value, str) and cell.column > 7:
        facilities.add(cell.value)
for f in sorted(facilities):
    print(f"  {f}")

# ── Other Sheets ────────────────────────────
print("\n" + "=" * 60)
print("OTHER SHEETS SUMMARY")
print("=" * 60)

for sheet_name in wb.sheetnames:
    if sheet_name == 'Lab':
        continue
    ws2 = wb[sheet_name]
    print(f"\n{sheet_name}: {ws2.max_row} rows x {ws2.max_column} cols")
    # Print first row (headers)
    headers = []
    for cell in next(ws2.iter_rows(max_row=1)):
        if cell.value:
            headers.append(f"{cell.column_letter}:{cell.value}")
    if headers:
        print(f"  Headers: {headers[:15]}")
        if len(headers) > 15:
            print(f"  ... +{len(headers) - 15} more")
    # Print row 2 sample
    try:
        row2 = list(ws2.iter_rows(min_row=2, max_row=2, values_only=False))
        if row2:
            vals = [(c.column_letter, c.value) for c in row2[0] if c.value is not None]
            print(f"  Row 2 sample: {vals[:10]}")
    except:
        pass

wb.close()
