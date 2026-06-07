#!/usr/bin/env python3
"""Examine 2025FlowSheets.xlsx and OneDrive 2023FlowSheets.xlsx."""
import openpyxl

ONEDRIVE = "/Users/woleakpose/Library/Group Containers/UBF8T346G9.OneDriveStandaloneSuite/OneDrive.noindex/OneDrive"

for label, path in [
    ("2025FlowSheets (OneDrive)", f"{ONEDRIVE}/2025FlowSheets.xlsx"),
    ("2023FlowSheets (OneDrive/6igmaHealth)", f"{ONEDRIVE}/6igmaHealth/2023FlowSheets.xlsx"),
]:
    print(f"\n{'='*60}")
    print(f"FILE: {label}")
    print(f"{'='*60}")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR: {e}")
        continue

    sheets = wb.sheetnames
    print(f"SHEETS ({len(sheets)}):")
    for i, s in enumerate(sheets):
        print(f"  {i+1}. {s}")

    skip = {'Sheet1','Sheet2','Sheet3','Summary','VomitLog','AccessMaintenance',
            'Supplies Request','Vomit Log','VomitLog2023','Access Maintenance',
            'Template','template','VomitLog2025','Vomit Log 2025'}
    session_sheets = [s for s in sheets if s not in skip]
    print(f"\nSESSION SHEETS: {len(session_sheets)}")
    if session_sheets:
        print(f"  First: {session_sheets[0]}")
        print(f"  Last:  {session_sheets[-1]}")

        # Examine first session
        ws = wb[session_sheets[0]]
        print(f"\n--- First session: '{session_sheets[0]}' ---")
        print(f"  Max rows: {ws.max_row}, Max cols: {ws.max_column}")
        for row in ws.iter_rows(max_row=55, values_only=False):
            vals = [(f"{c.column_letter}{c.row}", c.value) for c in row if c.value is not None]
            if vals:
                print(f"  {vals}")

    # Check non-session sheets
    non_session = [s for s in sheets if s in skip]
    if non_session:
        print(f"\nNON-SESSION SHEETS: {non_session}")
        for ns in non_session:
            ws = wb[ns]
            print(f"  '{ns}': rows={ws.max_row}")
    wb.close()
