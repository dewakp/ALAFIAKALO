#!/usr/bin/env python3
"""Examine a real 2025 session sheet with actual data."""
import openpyxl

ONEDRIVE = "/Users/woleakpose/Library/Group Containers/UBF8T346G9.OneDriveStandaloneSuite/OneDrive.noindex/OneDrive"
path = f"{ONEDRIVE}/2025FlowSheets.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

# Pick a real session (not Dummy)
ws = wb['01-03-2025']
print(f"=== SESSION: 01-03-2025 ===")
print(f"Max rows: {ws.max_row}, Max cols: {ws.max_column}")
print()
for row in ws.iter_rows(max_row=55, values_only=False):
    vals = [(f"{c.column_letter}{c.row}", c.value) for c in row if c.value is not None]
    if vals:
        print(f"  {vals}")

print("\n\n=== VOMIT LOG (first 10 rows) ===")
ws2 = wb['Vomit Log']
for i, row in enumerate(ws2.iter_rows(max_row=10, values_only=False)):
    vals = [(f"{c.column_letter}{c.row}", c.value) for c in row if c.value is not None]
    if vals:
        print(f"  Row {i+1}: {vals}")

wb.close()
