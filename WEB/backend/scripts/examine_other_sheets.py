#!/usr/bin/env python3
"""Examine Food, FoodLog, and Poop Log sheets."""
import openpyxl
from datetime import datetime

FILE = '/Users/woleakpose/Library/Mobile Documents/com~apple~CloudDocs/Records.xlsx'
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)

for sheet_name in ['Food', 'FoodLog', 'FoodLog2022', 'FoodLog2023', 'FoodLog2024', 'Poop Log', 'Recipe', 'Physicians']:
    if sheet_name not in wb.sheetnames:
        print(f"{sheet_name}: NOT FOUND")
        continue
    ws = wb[sheet_name]
    print(f"\n{'=' * 60}")
    print(f"SHEET: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols)")
    print('=' * 60)
    
    # Headers
    try:
        row1 = list(ws.iter_rows(max_row=1, values_only=False))
        if row1:
            headers = [(c.column_letter, c.value) for c in row1[0] if c.value is not None]
            print(f"  Headers: {headers[:20]}")
            if len(headers) > 20:
                print(f"  ... +{len(headers) - 20} more")
    except StopIteration:
        print("  (empty sheet)")
        continue
    
    # Sample rows
    print("  Sample rows:")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=False)):
        vals = [(c.column_letter, c.value) for c in row if c.value is not None]
        if vals:
            print(f"    Row {i+2}: {vals[:15]}")

wb.close()
