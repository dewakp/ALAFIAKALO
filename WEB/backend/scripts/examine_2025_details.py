#!/usr/bin/env python3
"""Check a few 2025 sessions for intradialytic reading data."""
import openpyxl

ONEDRIVE = "/Users/woleakpose/Library/Group Containers/UBF8T346G9.OneDriveStandaloneSuite/OneDrive.noindex/OneDrive"
path = f"{ONEDRIVE}/2025FlowSheets.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

# Check several sessions for intradialytic monitoring data (rows 31-44)
test_sheets = ['11-16-2025', '06-03-2025', '03-25-2025', '01-10-2025']
for sname in test_sheets:
    ws = wb[sname]
    print(f"\n=== {sname} - Intradialytic Readings (rows 29-44) ===")
    for row in ws.iter_rows(min_row=29, max_row=44, values_only=False):
        vals = [(f"{c.column_letter}{c.row}", c.value) for c in row if c.value is not None]
        if vals:
            print(f"  {vals}")
    
    # Also check post-treatment (rows 48-52)
    print(f"  --- Post Treatment ---")
    for row in ws.iter_rows(min_row=48, max_row=52, values_only=False):
        vals = [(f"{c.column_letter}{c.row}", c.value) for c in row if c.value is not None]
        if vals:
            print(f"  {vals}")
    
    # Check sitting BP values (row 11), standing BP (row 12)
    print(f"  --- Pre Vitals ---")
    for row in ws.iter_rows(min_row=11, max_row=17, values_only=False):
        vals = [(f"{c.column_letter}{c.row}", c.value) for c in row if c.value is not None]
        if vals:
            print(f"  {vals}")
    
    # Start/stop times
    for row in ws.iter_rows(min_row=24, max_row=24, values_only=False):
        vals = [(f"{c.column_letter}{c.row}", c.value) for c in row if c.value is not None]
        if vals:
            print(f"  Start/Stop: {vals}")

wb.close()
